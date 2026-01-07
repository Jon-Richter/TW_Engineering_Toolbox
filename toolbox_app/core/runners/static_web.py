from __future__ import annotations

import csv
import inspect
import json
import os
import socket
import threading
import time
import webbrowser
from datetime import datetime
from http.server import SimpleHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from typing import Any, Optional

from PySide6.QtCore import QEventLoop, QUrl, Qt
from PySide6.QtGui import QAction
from PySide6.QtWidgets import QApplication, QMainWindow, QMessageBox, QToolBar, QWidget, QVBoxLayout

try:
    from PySide6.QtWebEngineWidgets import QWebEngineView  # type: ignore
except Exception:  # pragma: no cover
    QWebEngineView = None  # type: ignore

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from toolbox_app.core.paths import user_data_dir
from toolbox_app.core.tool_base import ToolMeta


JS_CAPTURE = r"""
(() => {
  try {
    if (window.__TOOLBOX_RESULTS__ !== undefined) {
      return { ok: true, source: "window.__TOOLBOX_RESULTS__", data: window.__TOOLBOX_RESULTS__ };
    }
    if (typeof window.getToolboxResults === "function") {
      return { ok: true, source: "window.getToolboxResults()", data: window.getToolboxResults() };
    }

    const keys = ["toolbox_results", "tower_crane_foundation_results", "results", "tcf_results"];
    for (const k of keys) {
      const v = window.localStorage.getItem(k);
      if (v) {
        try { return { ok: true, source: `localStorage:${k}`, data: JSON.parse(v) }; }
        catch (e) { return { ok: true, source: `localStorage:${k}`, data: v }; }
      }
    }

    const el = document.querySelector("[data-toolbox-results]") || document.querySelector("#toolbox-results");
    if (el) {
      const txt = (el.textContent || "").trim();
      if (txt) {
        try { return { ok: true, source: "DOM element", data: JSON.parse(txt) }; }
        catch (e) { return { ok: true, source: "DOM element", data: txt }; }
      }
    }

    return {
      ok: false,
      reason: "No results exporter found. Add window.__TOOLBOX_RESULTS__ (recommended) or window.getToolboxResults() to enable capture/export."
    };
  } catch (e) {
    return { ok: false, reason: String(e) };
  }
})()
"""


def _guess_tool_dir() -> Path:
    core_root = Path(__file__).resolve().parents[1]
    for frame_info in inspect.stack():
        mod = inspect.getmodule(frame_info.frame)
        if mod is None:
            continue
        mod_file = getattr(mod, "__file__", None)
        if not mod_file:
            continue
        p = Path(mod_file).resolve()
        if core_root in p.parents:
            continue
        return p.parent
    return Path.cwd()


def _timestamp() -> str:
    return datetime.now().strftime("%Y%m%d_%H%M%S")


def _safe_local_output_dir(tool_id: str) -> Path:
    root = user_data_dir() / tool_id
    root.mkdir(parents=True, exist_ok=True)
    return root


def _is_scalar(v: Any) -> bool:
    return v is None or isinstance(v, (str, int, float, bool))


def _flatten(obj: Any, prefix: str = "") -> list[tuple[str, Any]]:
    out: list[tuple[str, Any]] = []
    if _is_scalar(obj):
        out.append((prefix or "value", obj))
        return out

    if isinstance(obj, dict):
        for k, v in obj.items():
            k_str = str(k)
            path = f"{prefix}.{k_str}" if prefix else k_str
            out.extend(_flatten(v, path))
        return out

    if isinstance(obj, (list, tuple)):
        for i, v in enumerate(obj):
            path = f"{prefix}[{i}]" if prefix else f"[{i}]"
            out.extend(_flatten(v, path))
        return out

    out.append((prefix or "value", str(obj)))
    return out


def _autosize(ws) -> None:
    widths: dict[int, int] = {}
    for row in ws.iter_rows(values_only=True):
        for i, cell in enumerate(row, start=1):
            if cell is None:
                continue
            widths[i] = max(widths.get(i, 0), len(str(cell)))
    for i, w in widths.items():
        ws.column_dimensions[get_column_letter(i)].width = min(max(w + 2, 12), 80)


def _export_excel(capture_payload: dict[str, Any], out_path: Path) -> None:
    wb = Workbook()

    ws = wb.active
    ws.title = "Summary"
    ws.append(["Path", "Value"])
    for path, value in _flatten(capture_payload):
        ws.append([path, value if _is_scalar(value) else json.dumps(value, ensure_ascii=False)])
    _autosize(ws)

    ws2 = wb.create_sheet("JSON")
    ws2.append(["JSON"])
    ws2.append([json.dumps(capture_payload, indent=2, ensure_ascii=False)])
    ws2.column_dimensions["A"].width = 120

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def _sanitize_identifier(name: str) -> str:
    s = "".join(c if c.isalnum() or c == "_" else "_" for c in name)
    while "__" in s:
        s = s.replace("__", "_")
    s = s.strip("_")
    if not s:
        s = "var"
    if s[0].isdigit():
        s = f"v_{s}"
    return s


def _export_mathcad_handoff(capture_payload: dict[str, Any], out_dir: Path) -> tuple[Path, Path, Path]:
    out_dir.mkdir(parents=True, exist_ok=True)

    json_path = out_dir / "handoff.json"
    csv_path = out_dir / "handoff.csv"
    txt_path = out_dir / "assignments.txt"

    json_path.write_text(json.dumps(capture_payload, indent=2, ensure_ascii=False), encoding="utf-8")

    flat = _flatten(capture_payload)
    with csv_path.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["path", "value"])
        for path, value in flat:
            if _is_scalar(value):
                w.writerow([path, value])
            else:
                w.writerow([path, json.dumps(value, ensure_ascii=False)])

    lines: list[str] = []
    lines.append(f"# Generated {datetime.now().isoformat(timespec='seconds')}")
    lines.append("# Note: only scalar values are emitted as direct assignments.")
    lines.append("")
    used: set[str] = set()
    for path, value in flat:
        if not _is_scalar(value):
            continue
        var = _sanitize_identifier(path)
        if var in used:
            continue
        used.add(var)
        if isinstance(value, str):
            rhs = f"\"{value}\""
        elif value is None:
            rhs = "NaN"
        elif isinstance(value, bool):
            rhs = "1" if value else "0"
        else:
            rhs = str(value)
        lines.append(f"{var} := {rhs}")

    txt_path.write_text("\n".join(lines) + "\n", encoding="utf-8")
    return json_path, csv_path, txt_path


class StaticWebRunner:
    RUNS_ON_UI_THREAD = True
    InputModel = None

    def __init__(
        self,
        meta: ToolMeta,
        index_file: str | Path,
        *,
        tool_dir: str | Path | None = None,
        window_title: str | None = None,
        open_in_browser: bool | None = None,
    ) -> None:
        self.meta = meta
        self._index_file = str(index_file)
        self._tool_dir = Path(tool_dir) if tool_dir is not None else _guess_tool_dir()
        self._window_title = window_title or meta.name
        self._open_in_browser = open_in_browser
        self._window: Optional[_StaticWebWindow] = None
        self._external: Optional[_StaticExternalLauncher] = None

    def _use_external(self) -> bool:
        if self._open_in_browser is True:
            return True
        if self._open_in_browser is False:
            return False
        return QWebEngineView is None

    def default_inputs(self) -> dict[str, Any]:
        return {}

    def run(self, inputs: dict[str, Any]) -> dict[str, Any]:
        if self._use_external():
            if self._external is None or not self._external.is_running():
                self._external = _StaticExternalLauncher(
                    tool_id=self.meta.id,
                    tool_name=self.meta.name,
                    tool_dir=self._tool_dir,
                    index_file=self._index_file,
                )
                self._external.start()
            return {"ok": True, "status": "launched", "mode": "external_browser"}

        if self._window is None or not self._window.isVisible():
            self._window = _StaticWebWindow(
                tool_id=self.meta.id,
                title=self._window_title,
                tool_dir=self._tool_dir,
                index_file=self._index_file,
            )
        self._window.show()
        self._window.raise_()
        self._window.activateWindow()
        return {"ok": True, "status": "launched", "mode": "embedded"}


class _StaticHandler(SimpleHTTPRequestHandler):
    def log_message(self, format: str, *args: Any) -> None:  # noqa: A002
        return


class _StaticAssetServer:
    def __init__(self, assets_dir: Path, index_rel: str) -> None:
        self.assets_dir = assets_dir
        self.index_rel = index_rel
        self._httpd: Optional[ThreadingHTTPServer] = None
        self._thread: Optional[threading.Thread] = None
        self.base_url: Optional[str] = None

    def start(self) -> None:
        if not self.assets_dir.exists():
            raise FileNotFoundError(f"Assets folder not found: {self.assets_dir}")

        sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        sock.bind(("127.0.0.1", 0))
        host, port = sock.getsockname()
        sock.close()

        def handler_factory(*args: Any, **kwargs: Any) -> _StaticHandler:
            return _StaticHandler(*args, directory=str(self.assets_dir), **kwargs)

        self._httpd = ThreadingHTTPServer((host, port), handler_factory)
        self.base_url = f"http://127.0.0.1:{port}/{self.index_rel}"

        self._thread = threading.Thread(target=self._httpd.serve_forever, daemon=True)
        self._thread.start()
        time.sleep(0.05)

    def stop(self) -> None:
        if self._httpd is not None:
            try:
                self._httpd.shutdown()
            except Exception:
                pass
        self._httpd = None
        self._thread = None
        self.base_url = None


class _StaticExternalLauncher:
    def __init__(self, *, tool_id: str, tool_name: str, tool_dir: Path, index_file: str) -> None:
        self._tool_id = tool_id
        self._tool_name = tool_name
        self._tool_dir = tool_dir
        self._index_file = index_file
        self._server: Optional[_StaticAssetServer] = None
        self._opened = False

        app = QApplication.instance()
        if app is None:
            raise RuntimeError("Qt application not initialized.")
        app.aboutToQuit.connect(self.stop)  # type: ignore[attr-defined]

    def is_running(self) -> bool:
        return bool(self._server and self._server.base_url)

    def start(self) -> None:
        index_path = Path(self._index_file)
        if not index_path.is_absolute():
            index_path = self._tool_dir / self._index_file
        index_path = index_path.resolve()

        if not index_path.exists():
            QMessageBox.critical(
                None,
                "Missing tool assets",
                f"Could not find index file at: {index_path}\n"
                "Ensure the static build output is present.",
            )
            return

        assets_dir = index_path.parent
        index_rel = index_path.relative_to(assets_dir).as_posix()

        self._server = _StaticAssetServer(assets_dir, index_rel)
        try:
            self._server.start()
        except Exception as e:
            QMessageBox.critical(None, "Failed to start local server", str(e))
            return

        if self._server.base_url and not self._opened:
            webbrowser.open(self._server.base_url)
            self._opened = True
            QMessageBox.information(
                None,
                f"{self._tool_name} (Browser Mode)",
                "Opened in your default browser.\n\n"
                "Capture/export from the Toolbox is not available in external browser mode.",
            )

    def stop(self) -> None:
        if self._server is not None:
            self._server.stop()
            self._server = None


class _StaticWebWindow(QMainWindow):
    def __init__(self, *, tool_id: str, title: str, tool_dir: Path, index_file: str) -> None:
        super().__init__()
        self.setWindowTitle(title)
        self.setMinimumSize(1100, 750)

        index_path = Path(index_file)
        if not index_path.is_absolute():
            index_path = tool_dir / index_file
        index_path = index_path.resolve()

        if not index_path.exists():
            QMessageBox.critical(
                self,
                "Missing tool assets",
                f"Could not find index file at: {index_path}\n"
                "Ensure the static build output is present.",
            )
            raise FileNotFoundError(str(index_path))

        assets_dir = index_path.parent
        index_rel = index_path.relative_to(assets_dir).as_posix()

        self._tool_id = tool_id
        self._server = _StaticAssetServer(assets_dir, index_rel)
        self._last_capture: Optional[dict[str, Any]] = None

        self._web = QWebEngineView()
        central = QWidget()
        lay = QVBoxLayout(central)
        lay.setContentsMargins(0, 0, 0, 0)
        lay.addWidget(self._web)
        self.setCentralWidget(central)

        self._make_toolbar()
        self._start()

    def closeEvent(self, event) -> None:  # type: ignore[override]
        try:
            self._server.stop()
        finally:
            return super().closeEvent(event)

    def _make_toolbar(self) -> None:
        tb = QToolBar("Actions")
        tb.setToolButtonStyle(Qt.ToolButtonStyle.ToolButtonTextBesideIcon)
        self.addToolBar(tb)

        act_capture = QAction("Capture results", self)
        act_capture.triggered.connect(self.capture_results)  # type: ignore[attr-defined]
        tb.addAction(act_capture)

        act_xlsx = QAction("Export to Excel (.xlsx)", self)
        act_xlsx.triggered.connect(self.export_to_excel)  # type: ignore[attr-defined]
        tb.addAction(act_xlsx)

        act_mathcad = QAction("Mathcad handoff (JSON/CSV/TXT)", self)
        act_mathcad.triggered.connect(self.export_to_mathcad)  # type: ignore[attr-defined]
        tb.addAction(act_mathcad)

        act_report = QAction("Save report HTML", self)
        act_report.triggered.connect(self.save_report_html)  # type: ignore[attr-defined]
        tb.addAction(act_report)

        tb.addSeparator()

        act_open = QAction("Open exports folder", self)
        act_open.triggered.connect(self.open_exports_folder)  # type: ignore[attr-defined]
        tb.addAction(act_open)

    def _start(self) -> None:
        try:
            self._server.start()
        except Exception as e:
            self._msg_error("Failed to start local server", str(e))
            return

        if not self._server.base_url:
            self._msg_error("Failed to start local server", "No URL generated.")
            return

        self._web.setUrl(QUrl(self._server.base_url))

    def _tool_out_dir(self) -> Path:
        return _safe_local_output_dir(self._tool_id)

    def _exports_dir(self) -> Path:
        p = self._tool_out_dir() / "exports"
        p.mkdir(parents=True, exist_ok=True)
        return p

    def capture_results(self) -> None:
        def _on_done(result: Any) -> None:
            if isinstance(result, dict) and result.get("ok") is True:
                payload = {
                    "captured_at": datetime.now().isoformat(timespec="seconds"),
                    "source": result.get("source", "unknown"),
                    "data": result.get("data"),
                }
                self._last_capture = payload
                out_path = self._exports_dir() / f"capture_{_timestamp()}.json"
                try:
                    out_path.write_text(json.dumps(payload, indent=2), encoding="utf-8")
                    self._msg_info(
                        "Capture successful",
                        f"Source: {payload['source']}\nSaved: {out_path}",
                    )
                except Exception as e:
                    self._msg_error("Capture succeeded but save failed", str(e))
            else:
                reason = ""
                if isinstance(result, dict):
                    reason = str(result.get("reason", "Unknown capture failure."))
                else:
                    reason = f"Unexpected capture return: {type(result)}"
                self._msg_error("Capture failed", reason)

        self._web.page().runJavaScript(JS_CAPTURE, _on_done)

    def export_to_excel(self) -> None:
        if not self._last_capture:
            self._msg_error("No captured results", "Click 'Capture results' first.")
            return

        out_path = self._exports_dir() / f"{self._tool_id}_{_timestamp()}.xlsx"
        try:
            _export_excel(self._last_capture, out_path)
            self._msg_info("Excel export complete", f"Saved: {out_path}")
        except Exception as e:
            self._msg_error("Excel export failed", str(e))

    def export_to_mathcad(self) -> None:
        if not self._last_capture:
            self._msg_error("No captured results", "Click 'Capture results' first.")
            return

        out_dir = self._exports_dir() / f"mathcad_{_timestamp()}"
        out_dir.mkdir(parents=True, exist_ok=True)
        try:
            paths = _export_mathcad_handoff(self._last_capture, out_dir)
            msg = "\n".join(str(p) for p in paths)
            self._msg_info("Mathcad handoff complete", f"Wrote:\n{msg}")
        except Exception as e:
            self._msg_error("Mathcad handoff failed", str(e))

    def save_report_html(self) -> None:
        out_path = self._exports_dir() / f"report_{_timestamp()}.html"

        loop = QEventLoop()
        html_holder: dict[str, str] = {}

        def _on_html(html: str) -> None:
            html_holder["html"] = html
            loop.quit()

        try:
            self._web.page().toHtml(_on_html)
            loop.exec()
            html = html_holder.get("html", "")
            if not html.strip():
                self._msg_error("Report export failed", "Empty HTML returned from WebEngine.")
                return
            out_path.write_text(html, encoding="utf-8")
            self._msg_info("Report saved", f"Saved: {out_path}")
        except Exception as e:
            self._msg_error("Report export failed", str(e))

    def open_exports_folder(self) -> None:
        p = self._exports_dir()
        try:
            if os.name == "nt":
                os.startfile(str(p))  # noqa: S606
            else:
                self._msg_info("Exports folder", str(p))
        except Exception as e:
            self._msg_error("Failed to open folder", str(e))

    def _msg_info(self, title: str, message: str) -> None:
        QMessageBox.information(self, title, message)

    def _msg_error(self, title: str, message: str) -> None:
        QMessageBox.critical(self, title, message)
