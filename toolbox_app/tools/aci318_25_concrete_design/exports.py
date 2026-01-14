\
from __future__ import annotations

import csv
import json
from pathlib import Path
from typing import Any, Dict, List, Tuple

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from calc_trace import CalcTrace
from report_renderer import render_report_html


def _autosize(ws):
    for col in range(1, ws.max_column + 1):
        max_len = 0
        for row in range(1, min(ws.max_row, 200) + 1):  # cap scanning
            v = ws.cell(row=row, column=col).value
            if v is None:
                continue
            max_len = max(max_len, len(str(v)))
        ws.column_dimensions[get_column_letter(col)].width = min(max(10, max_len + 2), 70)


def _write_json(path: Path, obj: Any) -> None:
    path.write_text(json.dumps(obj, indent=2, sort_keys=False), encoding="utf-8")


def export_all(trace: CalcTrace, results: Dict[str, Any], run_dir: Path) -> Dict[str, Any]:
    run_dir.mkdir(parents=True, exist_ok=True)

    meta = trace.meta
    vtag = f"v{meta.tool_version}_{meta.input_hash[:12]}"

    # 1) HTML report (authoritative)
    report_html = render_report_html(trace, results)
    (run_dir / "report.html").write_text(report_html, encoding="utf-8")
    # also hashed copy for audit trail
    (run_dir / f"report_{vtag}.html").write_text(report_html, encoding="utf-8")

    # 2) JSON exports
    trace_dict = trace.to_json_dict()
    _write_json(run_dir / "calc_trace.json", trace_dict)
    _write_json(run_dir / f"calc_trace_{vtag}.json", trace_dict)

    _write_json(run_dir / "results.json", results)
    _write_json(run_dir / f"results_{vtag}.json", results)

    # 3) Excel export
    wb = Workbook()
    # Inputs
    ws_in = wb.active
    ws_in.title = "Inputs"
    ws_in.append(["id", "label", "value", "units", "source", "notes"])
    for i in trace.inputs:
        ws_in.append([i.id, i.label, i.value, i.units, i.source, i.notes or ""])
    _autosize(ws_in)

    # Assumptions
    ws_a = wb.create_sheet("Assumptions")
    ws_a.append(["id", "text"])
    for a in trace.assumptions:
        ws_a.append([a.id, a.text])
    _autosize(ws_a)

    # Calcs
    ws_c = wb.create_sheet("Calcs")
    ws_c.append(["id", "section", "title", "reference", "equation", "substitution", "result_rounded", "units", "variables_json"])
    for st in trace.steps:
        refs = "; ".join([f"{r.type}:{r.ref}" for r in st.references])
        vars_json = json.dumps([v.model_dump(mode="json") for v in st.variables], ensure_ascii=False)
        ws_c.append([
            st.id,
            st.section,
            st.title,
            refs,
            st.equation_latex,
            st.substitution_latex,
            st.result_rounded.value,
            st.result_rounded.units,
            vars_json,
        ])
    _autosize(ws_c)

    # Tables
    ws_t = wb.create_sheet("Tables")
    ws_t.append(["table_name", "json"])
    for k, v in trace.tables.items():
        ws_t.append([k, json.dumps(v, ensure_ascii=False)])
    _autosize(ws_t)

    xlsx_path = run_dir / "results.xlsx"
    wb.save(xlsx_path)
    # hashed copy
    wb.save(run_dir / f"results_{vtag}.xlsx")

    # 4) Mathcad handoff
    # mathcad_inputs.csv — include inputs + key outputs
    mc_inputs_path = run_dir / "mathcad_inputs.csv"
    with mc_inputs_path.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["id", "label", "value", "units", "source"])
        for i in trace.inputs:
            w.writerow([i.id, i.label, i.value, i.units, i.source])
        # key outputs
        w.writerow([])
        w.writerow(["key", "value", "units", "source"])
        for k, v in (trace.summary.key_outputs or {}).items():
            if isinstance(v, dict) and "value" in v and "units" in v:
                w.writerow([k, v["value"], v["units"], "summary"])
            else:
                w.writerow([k, json.dumps(v), "", "summary"])
    # hashed copy
    (run_dir / f"mathcad_inputs_{vtag}.csv").write_text(mc_inputs_path.read_text(encoding="utf-8"), encoding="utf-8")

    mc_steps_path = run_dir / "mathcad_steps.json"
    _write_json(mc_steps_path, [s.model_dump(mode="json") for s in trace.steps])
    _write_json(run_dir / f"mathcad_steps_{vtag}.json", [s.model_dump(mode="json") for s in trace.steps])

    return {
        "run_dir": str(run_dir),
        "files": [
            "report.html",
            "calc_trace.json",
            "results.json",
            "results.xlsx",
            "mathcad_inputs.csv",
            "mathcad_steps.json",
        ],
        "files_hashed": [
            f"report_{vtag}.html",
            f"calc_trace_{vtag}.json",
            f"results_{vtag}.json",
            f"results_{vtag}.xlsx",
            f"mathcad_inputs_{vtag}.csv",
            f"mathcad_steps_{vtag}.json",
        ],
    }
