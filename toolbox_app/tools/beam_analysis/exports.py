from __future__ import annotations

import json
from pathlib import Path
from typing import Dict, Any, List

from openpyxl import Workbook
from openpyxl.utils import get_column_letter


def _autosize(ws):
    for col in range(1, ws.max_column + 1):
        max_len = 0
        col_letter = get_column_letter(col)
        for row in range(1, ws.max_row + 1):
            v = ws.cell(row=row, column=col).value
            if v is None:
                continue
            max_len = max(max_len, len(str(v)))
        ws.column_dimensions[col_letter].width = min(60, max(10, max_len + 2))


def export_results_json(results: Dict[str, Any], path: Path) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(results, indent=2), encoding="utf-8")
    return path


def export_mathcad_assignments(results: Dict[str, Any], path: Path) -> Path:
    d = results["diagrams"]
    labels = d.get("labels", {})
    lines: List[str] = []
    lines.append("// Beam Analysis Handoff (copy/paste or adapt for your template)")
    lines.append(f"// x [{labels.get('x','')}] | V [{labels.get('V','')}] | M [{labels.get('M','')}] | v [{labels.get('v','')}]")
    lines.append("")
    lines.append("x := [" + ", ".join(f"{v:.6g}" for v in d["x_user"]) + "]")
    lines.append("V := [" + ", ".join(f"{v:.6g}" for v in d["shear_user"]) + "]")
    lines.append("M := [" + ", ".join(f"{v:.6g}" for v in d["moment_user"]) + "]")
    lines.append("v := [" + ", ".join(f"{v:.6g}" for v in d["deflection_user"]) + "]")
    lines.append("")
    lines.append("// Reactions (per sign convention in results['sign_convention'])")
    for r in results.get("reactions", []):
        lines.append(f"// joint {r.get('joint')} @ x={r.get('x_user')} : {r.get('dof')} = {r.get('reaction_user')}")
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text("\n".join(lines), encoding="utf-8")
    return path


def export_results_excel(results: Dict[str, Any], path: Path) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()

    ws_in = wb.active
    ws_in.title = "Inputs"
    ws_in.append(["Field", "Value"])
    ws_in.append(["Tool", "beam_analysis"])
    ws_in.append(["Unit System", results.get("unit_system")])
    ws_in.append(["Sign convention", json.dumps(results.get("sign_convention", {}))])

    inputs = results.get("inputs", {})
    ws_in.append(["", ""])
    ws_in.append(["Spans", ""])
    ws_in.append(["index", "length", "E", "I"])
    for i, s in enumerate(inputs.get("spans", [])):
        ws_in.append([i, s.get("length"), s.get("E"), s.get("I")])

    ws_in.append(["", ""])
    ws_in.append(["Joints", ""])
    ws_in.append(["index", "restraint_v", "restraint_theta", "internal_hinge"])
    for i, j in enumerate(inputs.get("joints", [])):
        ws_in.append([i, j.get("restraint_v"), j.get("restraint_theta"), j.get("internal_hinge")])

    ws_in.append(["", ""])
    ws_in.append(["Distributed Loads (global)", ""])
    ws_in.append(["x_start", "x_end", "w_start", "w_end"])
    for dl in inputs.get("distributed_loads", []):
        ws_in.append([dl.get("x_start"), dl.get("x_end"), dl.get("w_start"), dl.get("w_end")])

    ws_in.append(["", ""])
    ws_in.append(["Point Loads (global)", ""])
    ws_in.append(["x", "P"])
    for pl in inputs.get("point_loads", []):
        ws_in.append([pl.get("x"), pl.get("P")])

    ws_in.append(["", ""])
    ws_in.append(["Point Moments (global)", ""])
    ws_in.append(["x", "M"])
    for pm in inputs.get("point_moments", []):
        ws_in.append([pm.get("x"), pm.get("M")])

    _autosize(ws_in)

    ws_n = wb.create_sheet("Mesh Nodes")
    m = results["mesh"]
    ws_n.append(["node", "x", "v", "theta_rad", "hinge_node"])
    for i, (x, v, t, h) in enumerate(zip(m["node_x_user"], m["node_v_user"], m["node_theta_rad"], m["hinge_nodes"])):
        ws_n.append([i, x, v, t, bool(h)])
    _autosize(ws_n)

    ws_r = wb.create_sheet("Reactions")
    ws_r.append(["joint", "x", "dof", "reaction"])
    for r in results.get("reactions", []):
        ws_r.append([r.get("joint"), r.get("x_user"), r.get("dof"), r.get("reaction_user")])
    _autosize(ws_r)

    ws_d = wb.create_sheet("Diagrams")
    labels = results.get("diagrams", {}).get("labels", {})
    ws_d.append([f"x ({labels.get('x','')})", f"V ({labels.get('V','')})", f"M ({labels.get('M','')})", f"v ({labels.get('v','')})"])
    d = results["diagrams"]
    for x, V, M, v in zip(d["x_user"], d["shear_user"], d["moment_user"], d["deflection_user"]):
        ws_d.append([x, V, M, v])
    _autosize(ws_d)

    wb.save(path)
    return path
