from __future__ import annotations

import csv
import json
import math
from datetime import datetime
from html import escape
from pathlib import Path
from typing import Any, Dict, List

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from .design_common import (
    factors_compression,
    factors_flexure,
    factors_shear,
    factors_tension_rupture,
    factors_tension_yield,
    ft_to_in,
)


def export_excel(run_dir: Path, results: Dict[str, Any], trace: List[str]) -> Path:
    wb = Workbook()

    # Inputs
    ws_in = wb.active
    ws_in.title = "Inputs"
    inputs = results.get("inputs", {})
    row = 1
    ws_in.cell(row=row, column=1, value="Field")
    ws_in.cell(row=row, column=2, value="Value")
    row += 1
    for k in sorted(inputs.keys()):
        ws_in.cell(row=row, column=1, value=k)
        ws_in.cell(row=row, column=2, value=str(inputs[k]))
        row += 1
    _autosize(ws_in, 1, 2)

    # Shape
    ws_sh = wb.create_sheet("Section")
    shape = results.get("shape", {})
    row = 1
    ws_sh.cell(row=row, column=1, value="Property")
    ws_sh.cell(row=row, column=2, value="Value")
    row += 1
    for k in sorted(shape.keys()):
        ws_sh.cell(row=row, column=1, value=k)
        ws_sh.cell(row=row, column=2, value=shape[k])
        row += 1
    _autosize(ws_sh, 1, 2)

    # Results
    ws_res = wb.create_sheet("Results")
    ws_res.cell(row=1, column=1, value="Check")
    ws_res.cell(row=1, column=2, value="Key")
    ws_res.cell(row=1, column=3, value="Value")
    r = 2
    checks = results.get("checks", {})
    def _cell_value(value: Any) -> Any:
        if value is None or isinstance(value, (str, int, float, bool)):
            return value
        return json.dumps(value, ensure_ascii=False)

    for check_name in ["flexure", "shear", "tension", "compression", "interaction"]:
        d = checks.get(check_name, {})
        for k in sorted(d.keys()):
            ws_res.cell(row=r, column=1, value=check_name)
            ws_res.cell(row=r, column=2, value=k)
            ws_res.cell(row=r, column=3, value=_cell_value(d[k]))
            r += 1
    _autosize(ws_res, 1, 3)

    # Trace
    ws_tr = wb.create_sheet("Trace")
    ws_tr.cell(row=1, column=1, value="Computation Trace")
    for i, line in enumerate(trace, start=2):
        ws_tr.cell(row=i, column=1, value=line)
    _autosize(ws_tr, 1, 1)

    out = run_dir / "aisc360_design_results.xlsx"
    wb.save(out)
    return out


def export_mathcad_handoff(run_dir: Path, results: Dict[str, Any]) -> Dict[str, Path]:
    """
    Produces:
      - JSON: nested data (inputs, shape, checks)
      - CSV: flattened key/value list
      - TXT: flat assignments for quick Mathcad paste
    """
    json_path = run_dir / "mathcad_handoff.json"
    json_path.write_text(json.dumps(results, indent=2), encoding="utf-8")

    csv_path = run_dir / "mathcad_handoff.csv"
    flat = _flatten(results)
    with csv_path.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["path", "value"])
        for path, value in flat:
            w.writerow([path, value])

    assignments = []
    inputs = results.get("inputs", {})
    for k in sorted(inputs.keys()):
        assignments.append(f"{k} := {inputs[k]}")
    shape = results.get("shape", {})
    for k in sorted(shape.keys()):
        assignments.append(f"shape_{k} := {shape[k]}")
    checks = results.get("checks", {})
    for chk, d in checks.items():
        for k in sorted(d.keys()):
            assignments.append(f"{chk}_{k} := {d[k]}")

    assignments_path = run_dir / "mathcad_assignments.txt"
    assignments_path.write_text("\n".join(assignments), encoding="utf-8")

    return {"json_path": json_path, "csv_path": csv_path, "assignments_path": assignments_path}


def _flatten(obj: Any, prefix: str = "") -> List[tuple[str, Any]]:
    out: List[tuple[str, Any]] = []
    if obj is None or isinstance(obj, (str, int, float, bool)):
        out.append((prefix or "value", obj))
        return out
    if isinstance(obj, dict):
        for k, v in obj.items():
            k_str = str(k)
            path = f"{prefix}.{k_str}" if prefix else k_str
            out.extend(_flatten(v, path))
        return out
    if isinstance(obj, (list, tuple)):
        for i, v in enumerate(obj):
            path = f"{prefix}[{i}]" if prefix else f"[{i}]"
            out.extend(_flatten(v, path))
        return out
    out.append((prefix or "value", str(obj)))
    return out


def _fmt(value: Any, digits: int = 3) -> str:
    if value is None:
        return "-"
    if isinstance(value, bool):
        return "1" if value else "0"
    if isinstance(value, (int, float)):
        if isinstance(value, float) and (math.isnan(value) or math.isinf(value)):
            return str(value)
        s = f"{float(value):.{digits}f}"
        s = s.rstrip("0").rstrip(".")
        return s
    return str(value)


def _fmt_unit(value: Any, unit: str, digits: int = 3) -> str:
    v = _fmt(value, digits)
    return v if v == "-" or unit == "" else f"{v} {unit}"


def _is_rect_hss(shape: Dict[str, Any]) -> bool:
    if str(shape.get("type_code", "")).strip().upper() != "HSS":
        return False
    label = str(shape.get("label", "")).upper().replace(" ", "")
    if not label.startswith("HSS"):
        return False
    core = label[3:]
    return core.count("X") != 1


def _h_tw_from_shape(shape: Dict[str, Any]) -> float | None:
    h_tw = shape.get("h_tw")
    if isinstance(h_tw, (int, float)) and h_tw > 0.0:
        return float(h_tw)
    h = shape.get("h_in")
    tw = shape.get("tw_in")
    if h and tw:
        return float(h) / float(tw)
    d = shape.get("d_in")
    tf = shape.get("tf_in")
    if d and tf and tw:
        h_val = float(d) - 2.0 * float(tf)
        if h_val > 0.0:
            return h_val / float(tw)
    return None


def _kc_from_h_tw(h_tw: float | None) -> float:
    if not h_tw or h_tw <= 0.0:
        return 0.76
    kc = 4.0 / math.sqrt(h_tw)
    return max(0.35, min(0.76, kc))


def _limit_state_rows(results: Dict[str, Any]) -> List[Dict[str, Any]]:
    checks = results.get("checks", {})
    shape = results.get("shape", {})
    tcode = str(shape.get("type_code", "")).strip().upper()
    is_hss_rect = _is_rect_hss(shape)
    rows: List[Dict[str, Any]] = []

    def add(name: str, demand: Any, capacity: Any, unity: Any, status: Any, units: str, ref: str) -> None:
        rows.append(
            {
                "name": name,
                "demand": demand,
                "capacity": capacity,
                "unity": unity,
                "status": status,
                "units": units,
                "ref": ref,
            }
        )

    flex = checks.get("flexure", {})
    if tcode == "L":
        add(
            "Flexure (major w)",
            flex.get("Mux_kft"),
            flex.get("M_design_x_kft"),
            flex.get("unity_x"),
            flex.get("status_x"),
            "kip-ft",
            "AISC 360-16 F10 (principal axes)",
        )
        add(
            "Flexure (minor z)",
            flex.get("Muy_kft"),
            flex.get("M_design_y_kft"),
            flex.get("unity_y"),
            flex.get("status_y"),
            "kip-ft",
            "AISC 360-16 F10 (principal axes)",
        )
    else:
        flex_ref = None
        major = flex.get("major_axis") if isinstance(flex, dict) else None
        if isinstance(major, dict) and major.get("section"):
            flex_ref = f"AISC 360-16 {major.get('section')}"
        elif str(shape.get("type_code", "")).strip().upper() == "HSS":
            flex_ref = "AISC 360-16 F7" if is_hss_rect else "AISC 360-16 F8"
        else:
            flex_ref = "AISC 360-16 F2/F3"
        add(
            "Flexure (major x)",
            flex.get("Mux_kft"),
            flex.get("M_design_x_kft"),
            flex.get("unity_x"),
            flex.get("status_x"),
            "kip-ft",
            flex_ref,
        )
        add(
            "Flexure (minor y)",
            flex.get("Muy_kft"),
            flex.get("M_design_y_kft"),
            flex.get("unity_y"),
            flex.get("status_y"),
            "kip-ft",
            "AISC 360-16 F6",
        )

    shear = checks.get("shear", {})
    shear_ref = "AISC 360-16 G2"
    if tcode in {"L", "2L", "WT", "MT", "ST"}:
        shear_ref = "AISC 360-16 G3"
    elif tcode == "HSS":
        shear_ref = "AISC 360-16 G4" if is_hss_rect else "AISC 360-16 G5"
    elif tcode == "PIPE":
        shear_ref = "AISC 360-16 G5"
    add(
        "Shear (major x)",
        shear.get("Vux_k"),
        shear.get("Vd_x_k"),
        shear.get("unity_x"),
        shear.get("status_x"),
        "k",
        shear_ref,
    )
    add(
        "Shear (minor y)",
        shear.get("Vuy_k"),
        shear.get("Vd_y_k"),
        shear.get("unity_y"),
        shear.get("status_y"),
        "k",
        shear_ref,
    )

    tension = checks.get("tension", {})
    add(
        "Tension",
        tension.get("Pu_tension_k"),
        tension.get("R_design_k"),
        tension.get("unity"),
        tension.get("status"),
        "k",
        "AISC 360-16 D2/D3",
    )

    comp = checks.get("compression", {})
    add(
        "Compression",
        comp.get("Pu_compression_k"),
        comp.get("P_design_k"),
        comp.get("unity"),
        comp.get("status"),
        "k",
        "AISC 360-16 E3/E4/E7",
    )

    inter = checks.get("interaction", {})
    inter_ref = "AISC 360-16 H1"
    if isinstance(inter, dict) and inter.get("equation") == "AISC 360-16 H2-1":
        inter_ref = "AISC 360-16 H2-1"
    add(
        "Interaction",
        inter.get("unity"),
        1.0,
        inter.get("unity"),
        inter.get("status"),
        "",
        inter_ref,
    )

    return rows


def build_report_html(results: Dict[str, Any], trace: List[str]) -> Dict[str, Any]:
    inputs = results.get("inputs", {})
    shape = results.get("shape", {})
    material = results.get("material", {})
    checks = results.get("checks", {})
    tcode = str(shape.get("type_code", "")).strip().upper()
    is_hss_rect = _is_rect_hss(shape)
    warnings = results.get("warnings", [])

    rows = _limit_state_rows(results)
    max_unity = 0.0
    adequate = True
    for r in rows:
        u = r.get("unity")
        if isinstance(u, (int, float)):
            if u > max_unity and math.isfinite(float(u)):
                max_unity = float(u)
            if u > 1.0 + 1e-9:
                adequate = False
        if isinstance(r.get("status"), str) and r["status"] == "NG":
            adequate = False

    status_class = "pass" if adequate else "fail"
    status_text = "Member is Adequate" if adequate else "Member is NOT Adequate"

    def row_html(r: Dict[str, Any]) -> str:
        return (
            "<tr>"
            f"<td>{escape(str(r['name']))}</td>"
            f"<td class='num'>{escape(_fmt_unit(r['demand'], r['units']))}</td>"
            f"<td class='num'>{escape(_fmt_unit(r['capacity'], r['units']))}</td>"
            f"<td class='num'>{escape(_fmt(r['unity'], 3))}</td>"
            f"<td class='status {('ok' if r.get('status') == 'OK' else 'ng')}'>{escape(str(r.get('status', '-')))}</td>"
            f"<td class='ref'>{escape(str(r.get('ref', '')))}</td>"
            "</tr>"
        )

    flex = checks.get("flexure", {})
    major = flex.get("major_axis", {})
    minor = flex.get("minor_axis", {})
    shear = checks.get("shear", {})
    tension = checks.get("tension", {})
    comp = checks.get("compression", {})
    inter = checks.get("interaction", {})

    method = inputs.get("design_method", "ASD")
    ff = factors_flexure(method)
    fs = factors_shear(method)
    ft_y = factors_tension_yield(method)
    ft_u = factors_tension_rupture(method)
    fc = factors_compression(method)

    Fy = material.get("Fy_ksi")
    Fu = material.get("Fu_ksi")
    E = material.get("E_ksi")

    Lb_ft = inputs.get("Lb_ft", 0.0)
    Lb_in = ft_to_in(float(Lb_ft or 0.0))
    Cb = inputs.get("Cb", 1.0) or 1.0

    calc_blocks: List[str] = []

    # Flexure
    if tcode == "L":
        theta = major.get("theta_deg")
        Mw = major.get("Mw_kft")
        Mz = major.get("Mz_kft")
        ltb_w = major.get("ltb_w", {}) if isinstance(major, dict) else {}
        leg_w = major.get("leg_local_buckling_w", {}) if isinstance(major, dict) else {}
        leg_z = major.get("leg_local_buckling_z", {}) if isinstance(major, dict) else {}
        rz = math.sqrt(shape.get("Iz_in4") / shape.get("A_in2")) if shape.get("A_in2") and shape.get("Iz_in4") else None
        t_leg = shape.get("t_in") or shape.get("t_des_in") or shape.get("t_nom_in")
        Cb_eff = min(Cb, 1.5) if Cb else None
        leg_w = major.get("leg_local_buckling_w", {}) if isinstance(major, dict) else {}
        leg_z = major.get("leg_local_buckling_z", {}) if isinstance(major, dict) else {}
        ltb_w = major.get("ltb_w", {}) if isinstance(major, dict) else {}
        My_w = major.get("My_w_kipin")
        Mn_leg_w = leg_w.get("Mn_kipin")
        Mn_leg_z = leg_z.get("Mn_kipin")
        Mn_ltb_w = ltb_w.get("Mn_kipin")
        Mcr = ltb_w.get("Mcr_kipin")
        beta_w = ltb_w.get("beta_w")
        ratio = (My_w / Mcr) if (My_w and Mcr) else None
        calc_lines = [
            "Single-angle flexure about principal axes (AISC 360-16 F10):",
            f"theta = atan(tan?) = {_fmt(theta)} deg",
            "Mw = Mx*cos(theta) + My*sin(theta)",
            "Mz = -Mx*sin(theta) + My*cos(theta)",
            f"Mx = {_fmt(flex.get('Mx_kft_raw'))} kip-ft, My = {_fmt(flex.get('My_kft_raw'))} kip-ft",
            f"Mw = {_fmt(Mw)} kip-ft, Mz = {_fmt(Mz)} kip-ft",
            f"Sw_min = {_fmt(major.get('Sw_min_in3'))} in^3, Sz_min = {_fmt(major.get('Sz_min_in3'))} in^3",
            f"My_w = Fy*Sw = {_fmt(Fy)}*{_fmt(major.get('Sw_min_in3'))} = {_fmt(My_w)} kip-in",
            f"My_z = Fy*Sz = {_fmt(Fy)}*{_fmt(major.get('Sz_min_in3'))} = {_fmt(Fy * (major.get('Sz_min_in3') or 0.0))} kip-in",
            f"Leg local buckling (w): {leg_w.get('case')}",
            f"lambda = {_fmt(leg_w.get('lambda'))}, lambda_p = {_fmt(leg_w.get('lambda_p'))}, lambda_r = {_fmt(leg_w.get('lambda_r'))}",
            f"Mn_leg_w = {_fmt(Mn_leg_w)} kip-in",
            f"Leg local buckling (z): {leg_z.get('case')}",
            f"lambda = {_fmt(leg_z.get('lambda'))}, lambda_p = {_fmt(leg_z.get('lambda_p'))}, lambda_r = {_fmt(leg_z.get('lambda_r'))}",
            f"Mn_leg_z = {_fmt(Mn_leg_z)} kip-in",
            f"LTB (w): {ltb_w.get('case')}",
            f"r_z = sqrt(Iz/A) = sqrt({_fmt(shape.get('Iz_in4'))}/{_fmt(shape.get('A_in2'))}) = {_fmt(rz)} in",
            f"beta_w = {_fmt(beta_w)}",
            "Mcr = (9*E*A*r_z*t*Cb_eff)/(8*Lb)*(sqrt(1+(4.4*beta_w*r_z/(Lb*t))^2) + 4.4*beta_w*r_z/(Lb*t))",
            f"Mcr = (9*{_fmt(E)}*{_fmt(shape.get('A_in2'))}*{_fmt(rz)}*{_fmt(t_leg)}*{_fmt(Cb_eff)})/(8*{_fmt(ltb_w.get('Lb_in'))})*(sqrt(1+(4.4*{_fmt(beta_w)}*{_fmt(rz)}/({_fmt(ltb_w.get('Lb_in'))}*{_fmt(t_leg)}))^2) + 4.4*{_fmt(beta_w)}*{_fmt(rz)}/({_fmt(ltb_w.get('Lb_in'))}*{_fmt(t_leg)})) = {_fmt(Mcr)} kip-in",
            f"My/Mcr = {_fmt(ratio)}",
            "Mn_ltb = (1.92 - 1.17*sqrt(My/Mcr))*My (F10-2) if My/Mcr<=1",
            "Mn_ltb = (0.92 - 0.17*(Mcr/My))*Mcr (F10-3) if My/Mcr>1",
            f"Mn_ltb = {_fmt(Mn_ltb_w)} kip-in",
            "Mn_w = min(1.5*My_w, Mn_leg_w, Mn_ltb)",
            f"Mn_w = min(1.5*{_fmt(My_w)}, {_fmt(Mn_leg_w)}, {_fmt(Mn_ltb_w)}) = {_fmt(major.get('Mn_w_kipin'))} kip-in",
            "Mn_z = min(1.5*My_z, Mn_leg_z)",
            f"Mn_z = min(1.5*{_fmt(Fy * (major.get('Sz_min_in3') or 0.0))}, {_fmt(Mn_leg_z)}) = {_fmt(major.get('Mn_z_kipin'))} kip-in",
            f"Md_w = {'phi*Mn_w' if method == 'LRFD' else 'Mn_w/omega'} = {_fmt(flex.get('M_design_x_kft'))} kip-ft",
            f"Md_z = {'phi*Mn_z' if method == 'LRFD' else 'Mn_z/omega'} = {_fmt(flex.get('M_design_y_kft'))} kip-ft",
            f"Unity w = |Mw|/Md_w = {_fmt(flex.get('unity_x'))}",
            f"Unity z = |Mz|/Md_z = {_fmt(flex.get('unity_y'))}",
        ]
        calc_blocks.append("<pre>" + escape("\n".join(calc_lines)) + "</pre>")
    elif tcode == "HSS" and not is_hss_rect:
        # Round HSS
        lam = major.get("lambda")
        lam_p = major.get("lambda_p")
        lam_r = major.get("lambda_r")
        Sx = shape.get("Sx_in3")
        Sy = shape.get("Sy_in3")
        Zx = shape.get("Zx_in3")
        Zy = shape.get("Zy_in3")
        case_x = str(major.get("case") or "")
        case_y = str(minor.get("case") or "")
        fcr_round = (0.33 * E / lam) if (lam and E) else None
        mn_x_line = ""
        mn_y_line = ""
        if "noncompact" in case_x.lower():
            mn_x_line = f"Mn_x = (0.021*E/(D/t)+Fy)*Sx = (0.021*{_fmt(E)}/{_fmt(lam)}+{_fmt(Fy)})*{_fmt(Sx)} = {_fmt(major.get('Mn_kipin'))} kip-in"
        elif "slender" in case_x.lower():
            mn_x_line = f"Fcr = 0.33*E/(D/t) = 0.33*{_fmt(E)}/{_fmt(lam)} = {_fmt(fcr_round)} ksi; Mn_x = Fcr*Sx = {_fmt(fcr_round)}*{_fmt(Sx)} = {_fmt(major.get('Mn_kipin'))} kip-in"
        else:
            mn_x_line = f"Mn_x = Fy*Zx = {_fmt(Fy)}*{_fmt(Zx)} = {_fmt(Fy * (Zx or 0.0))} kip-in"
        if "noncompact" in case_y.lower():
            mn_y_line = f"Mn_y = (0.021*E/(D/t)+Fy)*Sy = (0.021*{_fmt(E)}/{_fmt(lam)}+{_fmt(Fy)})*{_fmt(Sy)} = {_fmt(minor.get('Mn_kipin'))} kip-in"
        elif "slender" in case_y.lower():
            mn_y_line = f"Fcr = 0.33*E/(D/t) = 0.33*{_fmt(E)}/{_fmt(lam)} = {_fmt(fcr_round)} ksi; Mn_y = Fcr*Sy = {_fmt(fcr_round)}*{_fmt(Sy)} = {_fmt(minor.get('Mn_kipin'))} kip-in"
        else:
            mn_y_line = f"Mn_y = Fy*Zy = {_fmt(Fy)}*{_fmt(Zy)} = {_fmt(Fy * (Zy or 0.0))} kip-in"
        calc_lines = [
            "Round HSS flexure (AISC 360-16 F8):",
            f"D/t = {_fmt(lam)}, lambda_p = {_fmt(lam_p)}, lambda_r = {_fmt(lam_r)}",
            "Compact: Mn = Fy*Z (F8-1)",
            "Noncompact: Mn = (0.021*E/(D/t) + Fy)*S (F8-2)",
            "Slender: Fcr = 0.33*E/(D/t), Mn = Fcr*S (F8-3/F8-4)",
            f"Mn (x) = {_fmt(major.get('Mn_kipin'))} kip-in ({major.get('case')})",
            mn_x_line,
            f"Mn (y) = {_fmt(minor.get('Mn_kipin'))} kip-in ({minor.get('case')})",
            mn_y_line,
            f"Md_x = {'phi*Mn' if method == 'LRFD' else 'Mn/omega'} = {_fmt(flex.get('M_design_x_kft'))} kip-ft",
            f"Md_y = {'phi*Mn' if method == 'LRFD' else 'Mn/omega'} = {_fmt(flex.get('M_design_y_kft'))} kip-ft",
            f"Unity x = Mux/Md_x = {_fmt(flex.get('unity_x'))}",
            f"Unity y = Muy/Md_y = {_fmt(flex.get('unity_y'))}",
        ]
        calc_blocks.append("<pre>" + escape("\n".join(calc_lines)) + "</pre>")
    elif is_hss_rect:
        ltb = major.get("ltb", {}) if isinstance(major, dict) else {}
        B = shape.get("B_in") or shape.get("b_in")
        H = shape.get("H_in") or shape.get("h_in")
        t = shape.get("t_des_in") or shape.get("t_nom_in")
        Sx = shape.get("Sx_in3")
        Zx = shape.get("Zx_in3")
        G = E / (2.0 * (1.0 + 0.3)) if E else None
        Mcr = ltb.get("Mcr_kipin")
        lam_f = major.get("flange_lambda")
        lam_w = major.get("web_lambda")
        Mp = Fy * Zx if Fy and Zx else None
        Mn_f = None
        Mn_w = None
        if lam_f and B and t and Sx and Mp:
            if major.get("flange_class") == "noncompact":
                Mn_f = Mp - (Mp - Fy * Sx) * (3.57 * lam_f * math.sqrt(Fy / E) - 4.0)
            elif major.get("flange_class") == "slender":
                be = 1.92 * float(t) * math.sqrt(E / Fy) * (1.0 - 0.38 / lam_f * math.sqrt(E / Fy))
                be = min(be, float(B))
                Se = Sx * (be / float(B))
                Mn_f = Fy * Se
            else:
                Mn_f = Mp
        if lam_w and H and t and Sx and Mp:
            if major.get("web_class") == "noncompact":
                Mn_w = Mp - (Mp - Fy * Sx) * (0.305 * lam_w * math.sqrt(Fy / E) - 0.738)
            elif major.get("web_class") == "slender":
                be = 1.92 * float(t) * math.sqrt(E / Fy) * (1.0 - 0.38 / lam_w * math.sqrt(E / Fy))
                be = min(be, float(H))
                Se = Sx * (be / float(H))
                Mn_w = Fy * Se
            else:
                Mn_w = Mp
        calc_lines = [
            "Rectangular HSS flexure (AISC 360-16 F7):",
            f"Flange class (x) = {major.get('flange_class')}, Web class (x) = {major.get('web_class')}",
            f"lambda_f = {_fmt(lam_f)}, lambda_pf = {_fmt(major.get('flange_lambda_p'))}, lambda_rf = {_fmt(major.get('flange_lambda_r'))}",
            f"lambda_w = {_fmt(lam_w)}, lambda_pw = {_fmt(major.get('web_lambda_p'))}, lambda_rw = {_fmt(major.get('web_lambda_r'))}",
            f"Mp = Fy*Zx = {_fmt(Fy)}*{_fmt(Zx)} = {_fmt(Mp)} kip-in",
            "Flange noncompact: Mn = Mp - (Mp - Fy*S)*(3.57*lambda_f*sqrt(Fy/E) - 4.0)",
            f"Mn_f = {_fmt(Mn_f)} kip-in",
            "Web noncompact: Mn = Mp - (Mp - Fy*S)*(0.305*lambda_w*sqrt(Fy/E) - 0.738)",
            f"Mn_w = {_fmt(Mn_w)} kip-in",
            "LTB: Mcr = Cb*(pi/Lb)*sqrt(E*G*J*Iy), Mn_ltb = min(Mp, Mcr)",
            f"Mcr (x) = {_fmt(Mcr)} kip-in",
            f"Mn (x) = {_fmt(major.get('Mn_kipin'))} kip-in",
            f"Mn (y) = {_fmt(minor.get('Mn_kipin'))} kip-in",
            f"Md_x = {'phi*Mn' if method == 'LRFD' else 'Mn/omega'} = {_fmt(flex.get('M_design_x_kft'))} kip-ft",
            f"Md_y = {'phi*Mn' if method == 'LRFD' else 'Mn/omega'} = {_fmt(flex.get('M_design_y_kft'))} kip-ft",
            f"Unity x = Mux/Md_x = {_fmt(flex.get('unity_x'))}",
            f"Unity y = Muy/Md_y = {_fmt(flex.get('unity_y'))}",
        ]
        calc_blocks.append("<pre>" + escape("\n".join(calc_lines)) + "</pre>")
    else:
        section = major.get("section") if isinstance(major, dict) else None
        if section in {"F4", "F5"}:
            ltb = major.get("ltb", {}) if isinstance(major, dict) else {}
            cflb = major.get("flange_local_buckling", {}) if isinstance(major, dict) else {}
            cfy = major.get("compression_flange_yielding", {}) if isinstance(major, dict) else {}
            tfy = major.get("tension_flange_yielding", {}) if isinstance(major, dict) else {}
            Sxc = major.get("Sxc_in3")
            Sxt = major.get("Sxt_in3")
            Mp = major.get("Mp_kipin")
            Rpc = major.get("Rpc") or major.get("Rpg")
            rt = ltb.get("rt_in")
            Lp = ltb.get("Lp_in")
            Lr = ltb.get("Lr_in")
            Fcr_ltb = ltb.get("Fcr_ksi")
            Mn_ltb = ltb.get("Mn_kipin")
            Mn_cfy = cfy.get("Mn_kipin")
            Mn_cflb = cflb.get("Mn_kipin")
            Mn_tfy = tfy.get("Mn_kipin")
            Myc = (Fy * Sxc) if (Fy and Sxc) else None
            Myt = (Fy * Sxt) if (Fy and Sxt) else None
            lam_w = major.get("web_lambda")
            lam_pw = major.get("web_lambda_p")
            lam_rw = major.get("web_lambda_r")
            lam_f = None
            if shape.get("bf_in") and shape.get("tf_in"):
                lam_f = float(shape.get("bf_in")) / (2.0 * float(shape.get("tf_in")))
            lam_pf = 0.38 * math.sqrt(E / Fy) if Fy and E else None
            lam_rf = 1.00 * math.sqrt(E / Fy) if Fy and E else None
            Iy = shape.get("Iy_in4")
            bf = shape.get("bf_in")
            tf = shape.get("tf_in")
            Iyc_over_Iy = None
            if Iy and bf and tf:
                Iyc = float(tf) * float(bf) ** 3 / 12.0
                Iyc_over_Iy = Iyc / float(Iy)
            aw = None
            if bf and tf and shape.get("tw_in") and shape.get("d_in"):
                hc = float(shape.get("d_in")) / 2.0 - float(tf)
                aw = (hc * float(shape.get("tw_in"))) / (float(bf) * float(tf)) if hc > 0 else None
            aw_cap = min(aw, 10.0) if aw is not None else None
            Rpg = major.get("Rpg")

            calc_lines = [
                f"Major-axis flexure (AISC 360-16 {section}):",
                f"Mp = Fy * Zx = {_fmt(Fy)} * {_fmt(shape.get('Zx_in3'))} = {_fmt(major.get('Mp_kipin'))} kip-in",
                f"Sxc = {_fmt(Sxc)} in^3, Sxt = {_fmt(Sxt)} in^3",
                f"Myc = Fy*Sxc = {_fmt(Fy)}*{_fmt(Sxc)} = {_fmt(Myc)} kip-in",
                f"Myt = Fy*Sxt = {_fmt(Fy)}*{_fmt(Sxt)} = {_fmt(Myt)} kip-in",
                f"Lb = {_fmt(Lb_ft)} ft = {_fmt(Lb_in)} in",
                f"web lambda = {_fmt(lam_w)}, lambda_p = {_fmt(lam_pw)}, lambda_r = {_fmt(lam_rw)}",
            ]
            if section == "F4":
                calc_lines.extend(
                    [
                        f"Iyc/Iy = {_fmt(Iyc_over_Iy)}",
                        "Rpc = 1.0 if Iyc/Iy <= 0.23 else (Mp/Myc) or linear interpolation",
                        f"Rpc = {_fmt(Rpc)}",
                        f"CFY (F4-1): Mn_cfy = Rpc*Myc = {_fmt(Rpc)}*{_fmt(Myc)} = {_fmt(Mn_cfy)} kip-in",
                    ]
                )
                if "Lb<=Lp" in str(ltb.get("case")):
                    calc_lines.append("LTB (F4-2): Mn_ltb = Rpc*Myc")
                    calc_lines.append(f"Mn_ltb = {_fmt(Rpc)}*{_fmt(Myc)} = {_fmt(Mn_ltb)} kip-in")
                elif "Lp<Lb<=Lr" in str(ltb.get("case")):
                    calc_lines.append("LTB (F4-2): Mn_ltb = Cb*(Rpc*Myc - (Rpc*Myc - FL*Sxc)*((Lb-Lp)/(Lr-Lp)))")
                    calc_lines.append(f"Mn_ltb = {_fmt(Mn_ltb)} kip-in")
                else:
                    calc_lines.append("LTB (F4-3): Fcr = Cb*pi^2*E/(Lb/rt)^2*sqrt(1+0.078*(J/(Sxc*h0))*(Lb/rt)^2)")
                    calc_lines.append(f"Fcr = {_fmt(Fcr_ltb)} ksi; Mn_ltb = min(Fcr*Sxc, Rpc*Myc) = {_fmt(Mn_ltb)} kip-in")
                calc_lines.extend(
                    [
                        f"Lp = {_fmt(Lp)} in, Lr = {_fmt(Lr)} in, rt = {_fmt(rt)} in",
                        f"CFLB: {cflb.get('case')}",
                        f"lambda_f = {_fmt(lam_f)}, lambda_pf = {_fmt(lam_pf)}, lambda_rf = {_fmt(lam_rf)}",
                        f"Mn_cflb = {_fmt(Mn_cflb)} kip-in",
                        f"TFY: {tfy.get('case')} -> Mn_tfy = {_fmt(Mn_tfy)} kip-in",
                    ]
                )
            else:
                calc_lines.extend(
                    [
                        f"aw = hc*tw/(bfc*tfc) = {_fmt(aw)}, aw_cap = {_fmt(aw_cap)}",
                        "Rpg = 1 - (aw/(1200+300*aw))*(lambda_w - 5.7*sqrt(E/Fy))",
                        f"Rpg = {_fmt(Rpg)}",
                        f"CFY (F5-1): Mn_cfy = Rpg*Fy*Sxc = {_fmt(Rpg)}*{_fmt(Fy)}*{_fmt(Sxc)} = {_fmt(Mn_cfy)} kip-in",
                        f"LTB case: {ltb.get('case')}",
                        f"Fcr = {_fmt(Fcr_ltb)} ksi; Mn_ltb = Rpg*Fcr*Sxc = {_fmt(Rpg)}*{_fmt(Fcr_ltb)}*{_fmt(Sxc)} = {_fmt(Mn_ltb)} kip-in",
                        f"CFLB: {cflb.get('case')}",
                        f"lambda_f = {_fmt(lam_f)}, lambda_pf = {_fmt(lam_pf)}, lambda_rf = {_fmt(lam_rf)}",
                        f"Mn_cflb = {_fmt(Mn_cflb)} kip-in",
                        f"TFY: {tfy.get('case')} -> Mn_tfy = {_fmt(Mn_tfy)} kip-in",
                    ]
                )
            calc_lines.extend(
                [
                    f"Mn = min(limit states) = {_fmt(major.get('Mn_kipin'))} kip-in",
                    f"Md = {'phi*Mn' if method == 'LRFD' else 'Mn/omega'} = {_fmt(flex.get('M_design_x_kft'))} kip-ft",
                    f"Unity = Mux/Md = {_fmt(flex.get('unity_x'))}",
                ]
            )
            calc_blocks.append("<pre>" + escape("\n".join(calc_lines)) + "</pre>")
        elif section == "F9":
            ltb = major.get("ltb", {}) if isinstance(major, dict) else {}
            flb = major.get("flange_local_buckling", {}) if isinstance(major, dict) else {}
            wlb = major.get("web_leg_local_buckling", {}) if isinstance(major, dict) else {}
            Mp = major.get("Mp_kipin")
            Sx = shape.get("Sx_in3")
            My = Fy * Sx if Fy and Sx else None
            Lp = ltb.get("Lp_in")
            Lr = ltb.get("Lr_in")
            Mcr = ltb.get("Mcr_kipin")
            Mn_ltb = ltb.get("Mn_kipin")
            Mn_flb = flb.get("Mn_kipin")
            Mn_wlb = wlb.get("Mn_kipin")
            d = shape.get("d_in") or shape.get("b_in")
            Iy = shape.get("Iy_in4")
            J = shape.get("J_in4")
            B = None
            if Lb_in and Iy and J and d:
                B = 2.3 * (float(d) / float(Lb_in)) * math.sqrt(float(Iy) / float(J)) if J else None
            calc_lines = [
                "Major-axis flexure (AISC 360-16 F9):",
                f"Mp = Fy*Zx (limited) = {_fmt(Mp)} kip-in",
                f"My = Fy*Sx = {_fmt(Fy)}*{_fmt(Sx)} = {_fmt(My)} kip-in",
                f"Lp = 1.76*ry*sqrt(E/Fy) = {_fmt(Lp)} in",
                f"Lr = 1.95*(E/Fy)*(Iy/Sx)*sqrt(2.36*(E/Fy)*(Sx/J)+1) = {_fmt(Lr)} in",
                f"B = 2.3*(d/Lb)*sqrt(Iy/J) = {_fmt(B)}",
                f"Mcr = (1.95*E/Lb)*sqrt(Iy*J)*(B + sqrt(1+B^2)) = {_fmt(Mcr)} kip-in",
                f"LTB case: {ltb.get('case')}",
            ]
            if "Lp<Lb<=Lr" in str(ltb.get("case")):
                calc_lines.append("Mn_ltb = Mp - (Mp-My)*((Lb-Lp)/(Lr-Lp))")
                calc_lines.append(
                    f"Mn_ltb = {_fmt(Mp)} - ({_fmt(Mp)}-{_fmt(My)})*(({_fmt(Lb_in)}-{_fmt(Lp)})/({_fmt(Lr)}-{_fmt(Lp)})) = {_fmt(Mn_ltb)} kip-in"
                )
            elif "Lb>Lr" in str(ltb.get("case")):
                calc_lines.append("Mn_ltb = Mcr")
                calc_lines.append(f"Mn_ltb = {_fmt(Mcr)} kip-in")
            else:
                calc_lines.append("Mn_ltb = Mp")
                calc_lines.append(f"Mn_ltb = {_fmt(Mp)} kip-in")
            calc_lines.extend(
                [
                    f"FLB case: {flb.get('case')}",
                    f"Mn_flb = {_fmt(Mn_flb)} kip-in",
                    f"WLB case: {wlb.get('case')}",
                    f"Mn_wlb = {_fmt(Mn_wlb)} kip-in",
                    f"Mn = min(Mp, Mn_ltb, Mn_flb, Mn_wlb) = {_fmt(major.get('Mn_kipin'))} kip-in",
                    f"Md = {'phi*Mn' if method == 'LRFD' else 'Mn/omega'} = {_fmt(flex.get('M_design_x_kft'))} kip-ft",
                    f"Unity = Mux/Md = {_fmt(flex.get('unity_x'))}",
                ]
            )
            calc_blocks.append("<pre>" + escape("\n".join(calc_lines)) + "</pre>")
        else:
            ltb = major.get("ltb", {}) if isinstance(major, dict) else {}
            cflb = major.get("flange_local_buckling", {}) if isinstance(major, dict) else {}
            Mp = major.get("Mp_kipin")
            Sx = shape.get("Sx_in3")
            Lp = ltb.get("Lp_in")
            Lr = ltb.get("Lr_in")
            Fcr = ltb.get("Fcr_ksi")
            rts = shape.get("rts_in")
            J = shape.get("J_in4")
            h0 = shape.get("h0_in")
            ltb_case = str(ltb.get("case") or "")
            mn_ltb = ltb.get("Mn_kipin")
            ltb_lines = [f"LTB case: {ltb_case}"]
            if "Lb<=Lp" in ltb_case:
                ltb_lines.append(f"Mn_LTB = Mp = {_fmt(Mp)} kip-in")
            elif "Lp<Lb<=Lr" in ltb_case:
                ltb_lines.append(
                    "Mn_LTB = Cb*(Mp - (Mp - 0.7*Fy*Sx)*((Lb-Lp)/(Lr-Lp))) <= Mp"
                )
                ltb_lines.append(
                    f"Mn_LTB = {_fmt(Cb)}*({_fmt(Mp)} - ({_fmt(Mp)} - 0.7*{_fmt(Fy)}*{_fmt(Sx)})*(({_fmt(Lb_in)}-{_fmt(Lp)})/({_fmt(Lr)}-{_fmt(Lp)}))) = {_fmt(mn_ltb)} kip-in"
                )
            elif "Lb>Lr" in ltb_case:
                ltb_lines.append(
                    "Fcr = Cb*(pi^2*E/(Lb/rts)^2)*sqrt(1 + 0.078*(J/(Sx*h0))*(Lb/rts)^2)"
                )
                ltb_lines.append("Mn_LTB = min(Fcr*Sx, Mp)")
                ltb_lines.append(
                    f"Fcr = {_fmt(Cb)}*(pi^2*{_fmt(E)}/({_fmt(Lb_in)}/{_fmt(rts)})^2)*sqrt(1 + 0.078*({_fmt(J)}/({_fmt(Sx)}*{_fmt(h0)}))*({_fmt(Lb_in)}/{_fmt(rts)})^2) = {_fmt(Fcr)} ksi"
                )
                ltb_lines.append(
                    f"Mn_LTB = min({_fmt(Fcr)}*{_fmt(Sx)}, {_fmt(Mp)}) = {_fmt(mn_ltb)} kip-in"
                )
            ltb_lines.extend(
                [
                    f"Cb = {_fmt(Cb)}",
                    f"Lp = {_fmt(Lp)} in, Lr = {_fmt(Lr)} in",
                    f"rts = {_fmt(rts)} in, J = {_fmt(J)} in^4, h0 = {_fmt(h0)} in",
                    f"Fcr = {_fmt(Fcr)} ksi",
                    f"Mn_LTB = {_fmt(mn_ltb)} kip-in",
                ]
            )
            lam = cflb.get("lambda")
            lam_p = cflb.get("lambda_p")
            lam_r = cflb.get("lambda_r")
            Mn_cflb = cflb.get("Mn_kipin")
            h_tw = _h_tw_from_shape(shape)
            kc = _kc_from_h_tw(h_tw)
            cflb_lines = [f"Flange CFLB case: {cflb.get('case')}"]
            if lam is not None and lam_p is not None and lam_r is not None:
                cflb_lines.append(
                    f"lambda = {_fmt(lam)}, lambda_p = {_fmt(lam_p)}, lambda_r = {_fmt(lam_r)}"
                )
            if "Compact" in str(cflb.get("case", "")):
                cflb_lines.append(f"Mn_CFLB = Mp = {_fmt(Mp)} kip-in")
            elif "Noncompact" in str(cflb.get("case", "")):
                cflb_lines.append(
                    "Mn_CFLB = Mp - (Mp - 0.7*Fy*Sx)*((lambda-lambda_p)/(lambda_r-lambda_p))"
                )
                cflb_lines.append(
                    f"Mn_CFLB = {_fmt(Mp)} - ({_fmt(Mp)} - 0.7*{_fmt(Fy)}*{_fmt(Sx)})*(({_fmt(lam)}-{_fmt(lam_p)})/({_fmt(lam_r)}-{_fmt(lam_p)})) = {_fmt(Mn_cflb)} kip-in"
                )
            elif "Slender" in str(cflb.get("case", "")):
                cflb_lines.append("Mn_CFLB = 0.9*E*kc*Sx/lambda^2")
                cflb_lines.append(
                    f"kc = 4/sqrt(h/tw) = {_fmt(kc)}"
                )
                cflb_lines.append(
                    f"Mn_CFLB = 0.9*{_fmt(E)}*{_fmt(kc)}*{_fmt(Sx)}/({_fmt(lam)})^2 = {_fmt(Mn_cflb)} kip-in"
                )
            else:
                cflb_lines.append(f"Mn_CFLB = {_fmt(Mn_cflb)} kip-in")
            calc_lines = [
                "Major-axis flexure (AISC 360-16 F2/F3):",
                f"Mp = Fy * Zx = {_fmt(Fy)} * {_fmt(shape.get('Zx_in3'))} = {_fmt(major.get('Mp_kipin'))} kip-in",
                f"Lb = {_fmt(Lb_ft)} ft = {_fmt(Lb_in)} in",
                *ltb_lines,
                *cflb_lines,
                f"Web class: {major.get('web_class')}",
                f"Mn = min(Mn_LTB, Mn_CFLB) = {_fmt(major.get('Mn_kipin'))} kip-in",
                f"Md = {'phi*Mn' if method == 'LRFD' else 'Mn/omega'} = {_fmt(flex.get('M_design_x_kft'))} kip-ft",
                f"Unity = Mux/Md = {_fmt(flex.get('unity_x'))}",
            ]
            calc_blocks.append("<pre>" + escape("\n".join(calc_lines)) + "</pre>")

        # Flexure - minor axis
        minor_case = str(minor.get("case") or "")
        lam = minor.get("lambda")
        lam_p = minor.get("lambda_p")
        lam_r = minor.get("lambda_r")
        Sy = shape.get("Sy_in3")
        mp_minor = Fy * (shape.get("Zy_in3") or 0.0)
        fcr_minor = None
        if isinstance(lam, (int, float)) and lam and lam > 0.0:
            fcr_minor = 0.69 * (E / (lam**2))
        minor_lines = [f"Mn = {_fmt(minor.get('Mn_kipin'))} kip-in ({minor_case})"]
        if "noncompact" in minor_case.lower():
            minor_lines.insert(
                0,
                "Mn = Mp - (Mp - 0.7*Fy*Sy)*((lambda-lambda_p)/(lambda_r-lambda_p))",
            )
            minor_lines.insert(
                1,
                f"Mn = {_fmt(mp_minor)} - ({_fmt(mp_minor)} - 0.7*{_fmt(Fy)}*{_fmt(Sy)})*(({_fmt(lam)}-{_fmt(lam_p)})/({_fmt(lam_r)}-{_fmt(lam_p)})) = {_fmt(minor.get('Mn_kipin'))} kip-in",
            )
        elif "slender" in minor_case.lower():
            minor_lines.insert(0, "Fcr = 0.69*E/lambda^2, Mn = Fcr*Sy")
            minor_lines.insert(
                1,
                f"Fcr = 0.69*{_fmt(E)}/({_fmt(lam)})^2 = {_fmt(fcr_minor)} ksi; Mn = {_fmt(fcr_minor)}*{_fmt(Sy)} = {_fmt(minor.get('Mn_kipin'))} kip-in",
            )
        else:
            minor_lines.insert(0, "Mn = Mp = Fy*Zy")
            minor_lines.insert(
                1,
                f"Mn = {_fmt(Fy)}*{_fmt(shape.get('Zy_in3'))} = {_fmt(mp_minor)} kip-in",
            )
        minor_lines.extend(
            [
                f"Mp = Fy*Zy = {_fmt(Fy)}*{_fmt(shape.get('Zy_in3'))} = {_fmt(mp_minor)} kip-in",
                f"lambda = {_fmt(lam)}, lambda_p = {_fmt(lam_p)}, lambda_r = {_fmt(lam_r)}",
                f"Fcr = {_fmt(fcr_minor)} ksi",
            ]
        )
        calc_lines = [
            "Minor-axis flexure (AISC 360-16 F6):",
            *minor_lines,
            f"Md = {'phi*Mn' if method == 'LRFD' else 'Mn/omega'} = {_fmt(flex.get('M_design_y_kft'))} kip-ft  "
            f"(phi={_fmt(ff.phi)}, omega={_fmt(ff.omega)})",
            f"Unity = Muy/Md = {_fmt(flex.get('unity_y'))}",
        ]
        calc_blocks.append("<pre>" + escape("\n".join(calc_lines)) + "</pre>")

    # Shear
    shear_lines = ["Shear (AISC 360-16 Chapter G):"]
    if tcode in {"W", "S", "M", "HP", "C", "MC", "WT", "ST", "MT"}:
        h = shape.get("h_in")
        if not h and shape.get("d_in") and shape.get("tf_in"):
            h = float(shape.get("d_in")) - 2.0 * float(shape.get("tf_in"))
        tw = shape.get("tw_in")
        h_tw = _h_tw_from_shape(shape)
        kv = 5.34
        limit = 1.10 * math.sqrt(kv * E / Fy) if Fy and E else None
        shear_lines.append(f"Aw = h*tw = {_fmt(h)}*{_fmt(tw)} = {_fmt(shear.get('Avx_in2'))} in^2")
        shear_lines.append(f"h/tw = {_fmt(h_tw)}, limit = 1.10*sqrt(kv*E/Fy) = {_fmt(limit)} (kv={_fmt(kv)})")
        shear_lines.append(
            "Cv = 1.0 if h/tw <= limit else limit/(h/tw)"
        )
        shear_lines.append(
            f"Cv = {_fmt(shear.get('Cv_x'))}"
        )
    elif tcode == "HSS" and not shape.get("OD_in"):
        t = shape.get("t_des_in") or shape.get("t_nom_in")
        H = shape.get("h_in") or shape.get("H_in")
        B = shape.get("b_in") or shape.get("B_in")
        shear_lines.append(f"Avx = 2*t*H = 2*{_fmt(t)}*{_fmt(H)} = {_fmt(shear.get('Avx_in2'))} in^2")
        shear_lines.append(f"Avy = 2*t*B = 2*{_fmt(t)}*{_fmt(B)} = {_fmt(shear.get('Avy_in2'))} in^2")
        kv = 5.0
        hx = (float(H) / float(t)) if (H and t) else None
        hy = (float(B) / float(t)) if (B and t) else None
        limit_x = 1.10 * math.sqrt(kv * E / Fy) if Fy and E else None
        shear_lines.append(f"h/t (x) = {_fmt(hx)}, h/t (y) = {_fmt(hy)}, limit = {_fmt(limit_x)} (kv={_fmt(kv)})")
        shear_lines.append("Cv = 1.0 if h/t <= limit else limit/(h/t)")
        shear_lines.append(f"Cv_x = {_fmt(shear.get('Cv_x'))}, Cv_y = {_fmt(shear.get('Cv_y'))}")
    elif tcode in {"L", "2L"}:
        b = shape.get("b_in") or shape.get("B_in") or shape.get("d_in")
        t = shape.get("t_in") or shape.get("t_des_in") or shape.get("t_nom_in")
        h_tw = (float(b) / float(t)) if (b and t) else None
        kv = 1.2
        limit = 1.10 * math.sqrt(kv * E / Fy) if Fy and E else None
        shear_lines.append(f"Av = b*t = {_fmt(b)}*{_fmt(t)} = {_fmt(shear.get('Avx_in2'))} in^2")
        shear_lines.append(f"b/t = {_fmt(h_tw)}, limit = {_fmt(limit)} (kv={_fmt(kv)})")
        shear_lines.append("Cv = 1.0 if b/t <= limit else limit/(b/t)")
        shear_lines.append(f"Cv = {_fmt(shear.get('Cv_x'))}")
    else:
        shear_lines.append(f"Avx = {_fmt(shear.get('Avx_in2'))} in^2, Avy = {_fmt(shear.get('Avy_in2'))} in^2")
        shear_lines.append(f"Cv_x = {_fmt(shear.get('Cv_x'))}, Cv_y = {_fmt(shear.get('Cv_y'))}")

    shear_lines.extend(
        [
            f"Vn_x = 0.6*Fy*Avx*Cv = 0.6*{_fmt(Fy)}*{_fmt(shear.get('Avx_in2'))}*{_fmt(shear.get('Cv_x'))} = {_fmt(shear.get('Vn_x_k'))} k",
            f"Vn_y = 0.6*Fy*Avy*Cv = 0.6*{_fmt(Fy)}*{_fmt(shear.get('Avy_in2'))}*{_fmt(shear.get('Cv_y'))} = {_fmt(shear.get('Vn_y_k'))} k",
            f"Vd_x = {'phi*Vn' if method == 'LRFD' else 'Vn/omega'} = {_fmt(shear.get('Vd_x_k'))} k  "
            f"(phi={_fmt(fs.phi)}, omega={_fmt(fs.omega)})",
            f"Vd_y = {'phi*Vn' if method == 'LRFD' else 'Vn/omega'} = {_fmt(shear.get('Vd_y_k'))} k",
            f"Unity x = Vux/Vd_x = {_fmt(shear.get('unity_x'))}",
            f"Unity y = Vuy/Vd_y = {_fmt(shear.get('unity_y'))}",
        ]
    )
    calc_lines = shear_lines
    calc_blocks.append("<pre>" + escape("\n".join(calc_lines)) + "</pre>")

    # Tension
    Rn_y = tension.get("Rn_yield_k")
    Rn_u = tension.get("Rn_rupture_k")
    if method == "LRFD":
        Ry = (ft_y.phi or 0.0) * (Rn_y or 0.0)
        Ru = (ft_u.phi or 0.0) * (Rn_u or 0.0)
    else:
        Ry = (Rn_y or 0.0) / (ft_y.omega or 1.0)
        Ru = (Rn_u or 0.0) / (ft_u.omega or 1.0)
    calc_lines = [
        "Tension (AISC 360-16 D2/D3):",
        f"Rn_y = Fy*Ag = {_fmt(Fy)}*{_fmt(shape.get('A_in2'))} = {_fmt(Rn_y)} k",
        f"Rn_u = Fu*Ae = {_fmt(Fu)}*{_fmt(shape.get('A_in2'))} = {_fmt(Rn_u)} k",
        f"Ry = {'phi*Rn_y' if method == 'LRFD' else 'Rn_y/omega'} = {_fmt(ft_y.phi if method == 'LRFD' else 1.0)}*{_fmt(Rn_y)} = {_fmt(Ry)} k"
        if method == "LRFD"
        else f"Ry = Rn_y/omega = {_fmt(Rn_y)}/{_fmt(ft_y.omega)} = {_fmt(Ry)} k",
        f"Ru = {'phi*Rn_u' if method == 'LRFD' else 'Rn_u/omega'} = {_fmt(ft_u.phi if method == 'LRFD' else 1.0)}*{_fmt(Rn_u)} = {_fmt(Ru)} k"
        if method == "LRFD"
        else f"Ru = Rn_u/omega = {_fmt(Rn_u)}/{_fmt(ft_u.omega)} = {_fmt(Ru)} k",
        f"R = min(Ry, Ru) = {_fmt(tension.get('R_design_k'))} k",
        f"Unity = Pu_tension/R = {_fmt(tension.get('unity'))}",
    ]
    calc_blocks.append("<pre>" + escape("\n".join(calc_lines)) + "</pre>")

    # Compression
    klrx = comp.get("KLrx") or 0.0
    klry = comp.get("KLry") or 0.0
    e4 = comp.get("E4_details", {}) if isinstance(comp, dict) else {}
    e7 = comp.get("E7_details", []) if isinstance(comp, dict) else []
    ae = comp.get("Ae_in2") if isinstance(comp, dict) else None
    lam_r_e3 = 4.71 * math.sqrt(E / Fy) if Fy and E else None
    def _fe_ksi(klr: float) -> float | None:
        if not klr or klr <= 0.0:
            return None
        return (math.pi**2 * E) / (klr**2)
    def _fcr_from_fe(fe: float | None) -> float | None:
        if fe is None or fe <= 0.0:
            return None
        if (Fy / fe) <= 2.25:
            return (0.658 ** (Fy / fe)) * Fy
        return 0.877 * fe
    Fe_x = _fe_ksi(float(klrx)) if klrx else None
    Fe_y = _fe_ksi(float(klry)) if klry else None
    Fcr_x = _fcr_from_fe(Fe_x)
    Fcr_y = _fcr_from_fe(Fe_y)
    Lz_ft = inputs.get("Lz_ft", 0.0) or 0.0
    Kz = inputs.get("Kz", 1.0) or 1.0
    Lcz = ft_to_in(float(Lz_ft)) * float(Kz) if Lz_ft else 0.0
    G = (E / (2.0 * (1.0 + 0.3))) if E else None
    Fez = None
    if Lcz > 0.0 and shape.get("A_in2") and e4.get("ro_in") and (shape.get("Cw_in6") or 0.0) is not None and (shape.get("J_in4") or 0.0) is not None:
        Fez = ((math.pi**2 * E * (shape.get("Cw_in6") or 0.0)) / (Lcz**2) + (G or 0.0) * (shape.get("J_in4") or 0.0)) / (
            (shape.get("A_in2") or 0.0) * (e4.get("ro_in") or 1.0) ** 2
        )
    def _fcr_line(fe: float | None, label: str) -> str:
        if fe is None or fe <= 0.0:
            return f"Fcr_{label} = -"
        if (Fy / fe) <= 2.25:
            val = (0.658 ** (Fy / fe)) * Fy
            return f"Fcr_{label} = 0.658^(Fy/Fe_{label})*Fy = 0.658^({_fmt(Fy)}/{_fmt(fe)})*{_fmt(Fy)} = {_fmt(val)} ksi"
        val = 0.877 * fe
        return f"Fcr_{label} = 0.877*Fe_{label} = 0.877*{_fmt(fe)} = {_fmt(val)} ksi"

    fcrx_line = _fcr_line(Fe_x, "x")
    fcry_line = _fcr_line(Fe_y, "y")

    e4_eq = str(e4.get("equation", "") or "")
    e4_fe = e4.get("Fe_ksi")
    e4_fex = e4.get("Fex_ksi")
    e4_fey = e4.get("Fey_ksi")
    e4_fez = e4.get("Fez_ksi")
    e4_h = shape.get("H_const")
    fes = e4_fex if tcode in {"C", "MC"} else e4_fey
    e4_line = f"E4 Fe ({e4_eq}) = {_fmt(e4_fe)} ksi"
    if e4_eq == "E4-2":
        e4_line = (
            "Fe = ((pi^2*E*Cw)/Lcz^2 + G*J)/(Ix+Iy) = "
            f"(({_fmt(math.pi**2)}*{_fmt(E)}*{_fmt(shape.get('Cw_in6'))})/({_fmt(Lcz)})^2 + {_fmt(G)}*{_fmt(shape.get('J_in4'))})/"
            f"({_fmt(shape.get('Ix_in4'))}+{_fmt(shape.get('Iy_in4'))}) = {_fmt(e4_fe)} ksi"
        )
    elif e4_eq == "E4-3":
        e4_line = (
            "Fe = (Fes+Fez)/(2H)*(1 - sqrt(1 - (4*Fes*Fez*H/(Fes+Fez)^2))) = "
            f"({_fmt(fes)}+{_fmt(e4_fez)})/(2*{_fmt(e4_h)})*(1 - sqrt(1 - (4*{_fmt(fes)}*{_fmt(e4_fez)}*{_fmt(e4_h)}/({_fmt(fes)}+{_fmt(e4_fez)})^2))) = {_fmt(e4_fe)} ksi"
        )

    calc_lines = [
        "Compression (AISC 360-16 E3/E4/E7):",
        f"KL/rx = {_fmt(klrx)} ; KL/ry = {_fmt(klry)}",
        f"Fe_x = pi^2*E/(KL/rx)^2 = (pi^2*{_fmt(E)})/({_fmt(klrx)})^2 = {_fmt(Fe_x)} ksi",
        f"Fe_y = pi^2*E/(KL/ry)^2 = (pi^2*{_fmt(E)})/({_fmt(klry)})^2 = {_fmt(Fe_y)} ksi",
        f"lambda_r = 4.71*sqrt(E/Fy) = 4.71*sqrt({_fmt(E)}/{_fmt(Fy)}) = {_fmt(lam_r_e3)}",
        fcrx_line,
        fcry_line,
        f"r0 (calc) = {_fmt(e4.get('ro_in'))} in",
        f"Lcz = Kz*Lz = {_fmt(Kz)}*{_fmt(Lz_ft)} ft = {_fmt(Lcz)} in",
        f"Fez = ((pi^2*E*Cw)/Lcz^2 + G*J)/(A*r0^2) = ((pi^2*{_fmt(E)}*{_fmt(shape.get('Cw_in6'))})/({_fmt(Lcz)})^2 + {_fmt(G)}*{_fmt(shape.get('J_in4'))})/({_fmt(shape.get('A_in2'))}*{_fmt(e4.get('ro_in'))}^2) = {_fmt(Fez)} ksi",
        e4_line,
        f"Fcr = {_fmt(comp.get('Fcr_ksi'))} ksi",
        f"Ae = {_fmt(ae)} in^2 (Ag = {_fmt(shape.get('A_in2'))} in^2)",
        f"Pn = Fcr*Ae = {_fmt(comp.get('Fcr_ksi'))}*{_fmt(ae)} = {_fmt(comp.get('Pn_k'))} k",
        f"Pc = {'phi*Pn' if method == 'LRFD' else 'Pn/omega'} = {_fmt(comp.get('P_design_k'))} k  "
        f"(phi={_fmt(fc.phi)}, omega={_fmt(fc.omega)})",
        f"Unity = Pu_compression/Pc = {_fmt(comp.get('unity'))}",
    ]
    calc_blocks.append("<pre>" + escape("\n".join(calc_lines)) + "</pre>")
    if isinstance(e7, list) and e7:
        rows_html = []
        for d in e7:
            rows_html.append(
                "<tr>"
                f"<td>{escape(str(d.get('element')))}</td>"
                f"<td class='num'>{escape(_fmt(d.get('lambda')))}</td>"
                f"<td class='num'>{escape(_fmt(d.get('lambda_r')))}</td>"
                f"<td class='num'>{escape(_fmt(d.get('b_in')))}</td>"
                f"<td class='num'>{escape(_fmt(d.get('t_in')))}</td>"
                f"<td class='num'>{escape(_fmt(d.get('c1')))}</td>"
                f"<td class='num'>{escape(_fmt(d.get('c2')))}</td>"
                f"<td class='num'>{escape(_fmt(d.get('be_in')))}</td>"
                f"<td class='num'>{escape(_fmt(d.get('reduction_in2')))}</td>"
                "</tr>"
            )
        calc_blocks.append(
            "<table>"
            "<thead><tr><th>Element</th><th>lambda</th><th>lambda_r</th><th>b</th><th>t</th><th>c1</th><th>c2</th><th>be</th><th>Area Red.</th></tr></thead>"
            "<tbody>"
            + "".join(rows_html)
            + "</tbody></table>"
        )
        formula_lines = [
            "E7 effective width (slender elements):",
            "Fel = (c2*lambda_r/lambda)^2 * Fy",
            "be = b * (1 - c1*sqrt(Fel/Fcr)) * sqrt(Fel/Fcr)",
            "Ae = Ag - sum((b-be)*t) over slender elements",
        ]
        for d in e7:
            lam = d.get("lambda")
            lam_r = d.get("lambda_r")
            c1 = d.get("c1")
            c2 = d.get("c2")
            b = d.get("b_in")
            t = d.get("t_in")
            Fel = None
            if lam and lam_r and c2 and Fy:
                Fel = (float(c2) * float(lam_r) / float(lam)) ** 2 * float(Fy)
            ratio = math.sqrt(max((Fel or 0.0) / (comp.get("Fcr_ksi") or 1.0), 0.0)) if Fel and comp.get("Fcr_ksi") else None
            be = d.get("be_in")
            formula_lines.extend(
                [
                    f"{d.get('element')}: Fel = ({_fmt(c2)}*{_fmt(lam_r)}/{_fmt(lam)})^2*{_fmt(Fy)} = {_fmt(Fel)} ksi",
                    f"{d.get('element')}: be = {_fmt(b)}*(1 - {_fmt(c1)}*sqrt(Fel/Fcr))*sqrt(Fel/Fcr) = {_fmt(be)} in",
                    f"{d.get('element')}: reduction = (b-be)*t = ({_fmt(b)}-{_fmt(be)})*{_fmt(t)} = {_fmt(d.get('reduction_in2'))} in^2",
                ]
            )
        calc_blocks.append("<pre>" + escape("\n".join(formula_lines)) + "</pre>")

    # Interaction
    if isinstance(inter, dict) and inter.get("equation") == "AISC 360-16 H2-1":
        calc_lines = [
            "Interaction (AISC 360-16 H2-1):",
            "f = P/A + Mw/Sw + Mz/Sz (evaluate at heel/toes)",
            "U = |fa|/Fa + |fbw|/Fbw + |fbz|/Fbz",
            "Points A/B/C are the AISC DB extreme points (two toes and the heel).",
            f"theta = {_fmt(inter.get('theta_deg'))} deg; Mw={_fmt(inter.get('Mw_kft'))} kip-ft, Mz={_fmt(inter.get('Mz_kft'))} kip-ft",
            f"Fa = {_fmt(inter.get('Fa_ksi'))} ksi",
            f"Md_w = {_fmt(flex.get('M_design_x_kft'))} kip-ft; Md_z = {_fmt(flex.get('M_design_y_kft'))} kip-ft",
            "Fbw = Md_w/Sw, Fbz = Md_z/Sz (point-specific)",
            f"Unity (max over points) = {_fmt(inter.get('unity'))}",
        ]
        calc_blocks.append("<pre>" + escape("\n".join(calc_lines)) + "</pre>")
        points = inter.get("points") if isinstance(inter, dict) else None
        if isinstance(points, list) and points:
            point_lines = []
            for p in points:
                point_lines.extend(
                    [
                        f"Point {p.get('point')}: fa = Pu/A = {_fmt(p.get('fa_ksi'))} ksi",
                        f"fbw = Mw*12/Sw = {_fmt(p.get('fbw_ksi'))} ksi; fbz = Mz*12/Sz = {_fmt(p.get('fbz_ksi'))} ksi",
                        f"U = |fa|/Fa + |fbw|/Fbw + |fbz|/Fbz = {_fmt(p.get('unity'))}",
                    ]
                )
            calc_blocks.append("<pre>" + escape("\n".join(point_lines)) + "</pre>")
            rows_html = []
            for p in points:
                rows_html.append(
                    "<tr>"
                    f"<td>{escape(str(p.get('point')))}</td>"
                    f"<td class='num'>{escape(_fmt(p.get('w_in')))}</td>"
                    f"<td class='num'>{escape(_fmt(p.get('z_in')))}</td>"
                    f"<td class='num'>{escape(_fmt(p.get('Sw_in3')))}</td>"
                    f"<td class='num'>{escape(_fmt(p.get('Sz_in3')))}</td>"
                    f"<td class='num'>{escape(_fmt(p.get('fa_ksi')))}</td>"
                    f"<td class='num'>{escape(_fmt(p.get('fbw_ksi')))}</td>"
                    f"<td class='num'>{escape(_fmt(p.get('fbz_ksi')))}</td>"
                    f"<td class='num'>{escape(_fmt(p.get('Fbw_ksi')))}</td>"
                    f"<td class='num'>{escape(_fmt(p.get('Fbz_ksi')))}</td>"
                    f"<td class='num'>{escape(_fmt(p.get('sigma_ksi')))}</td>"
                    f"<td class='num'>{escape(_fmt(p.get('unity')))}</td>"
                    f"<td class='status {('ok' if p.get('status') == 'OK' else 'ng')}'>{escape(str(p.get('status')))}</td>"
                    "</tr>"
                )
            calc_blocks.append(
                "<table>"
                "<thead><tr><th>Point</th><th>w</th><th>z</th><th>Sw</th><th>Sz</th><th>fa</th><th>fbw</th><th>fbz</th><th>Fbw</th><th>Fbz</th><th>sigma</th><th>unity</th><th>Status</th></tr></thead>"
                "<tbody>"
                + "".join(rows_html)
                + "</tbody></table>"
            )
    else:
        calc_lines = [
            "Interaction (AISC 360-16 H1):",
            f"Equation: {inter.get('equation')}",
            f"mx = Mux/Mcx = {_fmt(flex.get('Mux_kft'))}/{_fmt(flex.get('M_design_x_kft'))} = {_fmt(inter.get('mx'))}",
            f"my = Muy/Mcy = {_fmt(flex.get('Muy_kft'))}/{_fmt(flex.get('M_design_y_kft'))} = {_fmt(inter.get('my'))}",
        ]
        if inter.get("Pu_k") is not None and inter.get("Pc_k") is not None:
            pr = (float(inter.get("Pu_k") or 0.0) / float(inter.get("Pc_k") or 1.0)) if inter.get("Pc_k") else 0.0
            calc_lines.append(f"Pu/Pc = {_fmt(inter.get('Pu_k'))}/{_fmt(inter.get('Pc_k'))} = {_fmt(pr)}")
            if "8/9" in str(inter.get("equation", "")):
                calc_lines.append(f"U = Pu/Pc + (8/9)*(mx+my) = {_fmt(pr)} + (8/9)*({_fmt(inter.get('mx'))}+{_fmt(inter.get('my'))}) = {_fmt(inter.get('unity'))}")
            else:
                calc_lines.append(f"U = Pu/(2Pc) + (mx+my) = {_fmt(pr)}/2 + {_fmt(inter.get('mx'))}+{_fmt(inter.get('my'))} = {_fmt(inter.get('unity'))}")
        elif inter.get("Pt_k") is not None and inter.get("Pt_cap_k") is not None:
            tr = (float(inter.get("Pt_k") or 0.0) / float(inter.get("Pt_cap_k") or 1.0)) if inter.get("Pt_cap_k") else 0.0
            calc_lines.append(f"Pt/Pt_cap = {_fmt(inter.get('Pt_k'))}/{_fmt(inter.get('Pt_cap_k'))} = {_fmt(tr)}")
            calc_lines.append(f"U = Pt/Pt_cap + mx + my = {_fmt(tr)} + {_fmt(inter.get('mx'))}+{_fmt(inter.get('my'))} = {_fmt(inter.get('unity'))}")
        calc_lines.append(f"Unity = {_fmt(inter.get('unity'))}")
        calc_blocks.append("<pre>" + escape("\n".join(calc_lines)) + "</pre>")

    def table_from_dict(title: str, data: Dict[str, Any], columns: int = 2) -> str:
        rows = []
        items = sorted(data.items())
        if columns < 1:
            columns = 1
        for i in range(0, len(items), columns):
            chunk = items[i : i + columns]
            cells = []
            for k, v in chunk:
                cells.append(f"<td>{escape(str(k))}</td>")
                cells.append(f"<td class='num'>{escape(_fmt(v, 4))}</td>")
            for _ in range(columns - len(chunk)):
                cells.append("<td></td><td></td>")
            rows.append("<tr>" + "".join(cells) + "</tr>")
        header = "".join("<th>Field</th><th>Value</th>" for _ in range(columns))
        return (
            f"<h3>{escape(title)}</h3>"
            "<table class='kv'>"
            f"<thead><tr>{header}</tr></thead>"
            "<tbody>"
            + "".join(rows)
            + "</tbody></table>"
        )

    html = f"""
<!doctype html>
<html>
<head>
<meta charset="utf-8"/>
<title>AISC 360-16 Member Design Report</title>
<style>
body {{ font-family: "Segoe UI", Arial, sans-serif; color: #0f172a; }}
h1 {{ font-size: 20px; margin-bottom: 6px; }}
h2 {{ font-size: 16px; margin-top: 18px; }}
h3 {{ font-size: 14px; margin-top: 14px; }}
.summary {{ padding: 10px 12px; border-radius: 8px; margin: 8px 0 14px; }}
.summary.pass {{ background: #e6f4ea; border: 1px solid #2e7d32; color: #1b5e20; }}
.summary.fail {{ background: #fdecea; border: 1px solid #c62828; color: #7f1d1d; }}
.meta {{ font-size: 12px; color: #475569; }}
table {{ border-collapse: collapse; width: 100%; margin: 8px 0 16px; }}
th, td {{ border: 1px solid #d1d5db; padding: 6px 8px; font-size: 12px; }}
th {{ background: #f8fafc; text-align: left; }}
td.num {{ text-align: right; font-family: Consolas, monospace; }}
td.status.ok {{ color: #166534; font-weight: 600; }}
td.status.ng {{ color: #b91c1c; font-weight: 600; }}
td.ref {{ color: #475569; font-size: 11px; }}
pre {{ background: #f8fafc; padding: 10px; border-radius: 8px; border: 1px solid #e2e8f0; font-size: 12px; overflow: auto; }}
ul {{ margin: 6px 0 10px 18px; }}
</style>
</head>
<body>
<h1>AISC 360-16 Member Design Report</h1>
<div class="meta">Generated {escape(datetime.now().isoformat(timespec="seconds"))}</div>
<div class="summary {status_class}"><strong>{escape(status_text)}</strong> (max unity = {escape(_fmt(max_unity, 3))})</div>

<h2>Limit State Summary</h2>
<table>
<thead>
<tr>
<th>Limit State</th><th>Demand</th><th>Capacity</th><th>Unity</th><th>Status</th><th>Reference</th>
</tr>
</thead>
<tbody>
{''.join(row_html(r) for r in rows)}
</tbody>
</table>

<h2>Inputs</h2>
{table_from_dict("Design Inputs", {k: v for k, v in inputs.items() if k != "assume_no_holes"}, columns=3)}

<h2>Section Properties</h2>
{table_from_dict("Shape Properties", shape, columns=4)}

<h2>Step-by-Step Calculations</h2>
{''.join(calc_blocks)}

<h2>Warnings / Notes</h2>
<ul>
{''.join(f"<li>{escape(str(w))}</li>" for w in warnings) if warnings else "<li>None</li>"}
</ul>

<h2>Computation Trace</h2>
<pre>{escape(chr(10).join(trace))}</pre>
</body>
</html>
"""

    summary = {
        "adequate": adequate,
        "max_unity": max_unity,
        "status": "OK" if adequate else "NG",
    }
    return {"html": html, "summary": summary, "limit_states": rows}


def export_report_html(run_dir: Path, html: str) -> Path:
    out = run_dir / "aisc360_report.html"
    out.write_text(html, encoding="utf-8")
    return out


def _autosize(ws, col_min: int, col_max: int) -> None:
    for c in range(col_min, col_max + 1):
        max_len = 0
        for cell in ws[get_column_letter(c)]:
            v = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(v))
        ws.column_dimensions[get_column_letter(c)].width = min(70, max(12, max_len + 2))
