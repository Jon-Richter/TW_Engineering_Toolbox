from __future__ import annotations
from pathlib import Path
from typing import Any, Dict, List, Tuple
import json
from datetime import datetime

from openpyxl import Workbook
from openpyxl.utils import get_column_letter


def _flatten(prefix: str, obj: Any, out: List[Tuple[str, Any]]) -> None:
    if isinstance(obj, dict):
        for k, v in obj.items():
            _flatten(f"{prefix}{k}.", v, out)
    elif isinstance(obj, list):
        for i, v in enumerate(obj):
            _flatten(f"{prefix}{i}.", v, out)
    else:
        key = prefix[:-1] if prefix.endswith(".") else prefix
        out.append((key, obj))


def export_to_excel(path: Path, payload: Dict[str, Any]) -> Path:
    """
    Creates a straightforward workbook:
      - Summary: metadata
      - Inputs: key/value
      - Results: flattened key/value from lastResult
    """
    wb = Workbook()
    ws_sum = wb.active
    ws_sum.title = "Summary"
    ws_in = wb.create_sheet("Inputs")
    ws_out = wb.create_sheet("Results")

    ts = datetime.now().isoformat(timespec="seconds")
    ws_sum["A1"] = "Tool"
    ws_sum["B1"] = "Base Plate Design (Embedded HTML)"
    ws_sum["A2"] = "Timestamp"
    ws_sum["B2"] = ts
    ws_sum["A3"] = "baseType"
    ws_sum["B3"] = payload.get("baseType")
    ws_sum["A4"] = "ropeSize"
    ws_sum["B4"] = payload.get("ropeSize")

    # Inputs
    ws_in["A1"] = "Input"
    ws_in["B1"] = "Value"
    inputs = payload.get("inputs", {}) or {}
    for r, (k, v) in enumerate(sorted(inputs.items()), start=2):
        ws_in[f"A{r}"] = k
        ws_in[f"B{r}"] = v

    # Results
    ws_out["A1"] = "Result"
    ws_out["B1"] = "Value"
    rows: List[Tuple[str, Any]] = []
    _flatten("", payload.get("lastResult"), rows)
    for r, (k, v) in enumerate(rows, start=2):
        ws_out[f"A{r}"] = k
        ws_out[f"B{r}"] = "" if v is None else v

    # Basic formatting
    for ws in (ws_in, ws_out, ws_sum):
        ws.freeze_panes = "A2"
        ws.column_dimensions["A"].width = 34
        ws.column_dimensions["B"].width = 24

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return path


def export_mathcad_handoff(folder: Path, payload: Dict[str, Any]) -> List[Path]:
    """
    Creates simple handoff artifacts that Mathcad Prime can ingest (via file read or Excel component):
      - payload.json: full captured structure
      - inputs.csv: key,value
      - results.csv: flattened key,value
      - assignments.txt: key:=value lines for paste/import workflows

    This avoids tight coupling to Mathcad's automation model and is robust under read-only SharePoint.
    """
    folder.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")

    p_json = folder / f"catenary_guying_payload_{ts}.json"
    p_in = folder / f"catenary_guying_inputs_{ts}.csv"
    p_out = folder / f"catenary_guying_results_{ts}.csv"
    p_txt = folder / f"catenary_guying_assignments_{ts}.txt"

    p_json.write_text(json.dumps(payload, indent=2), encoding="utf-8")

    inputs = payload.get("inputs", {}) or {}
    p_in.write_text("name,value\n" + "\n".join([f"{k},{inputs[k]}" for k in sorted(inputs.keys())]) + "\n", encoding="utf-8")

    rows: List[Tuple[str, Any]] = []
    _flatten("", payload.get("lastResult"), rows)
    out_lines = ["name,value"]
    for k, v in rows:
        out_lines.append(f"{k},{'' if v is None else v}")
    p_out.write_text("\n".join(out_lines) + "\n", encoding="utf-8")

    # Assignments (for manual paste or text import)
    assign_lines = []
    for k in sorted(inputs.keys()):
        assign_lines.append(f"{k}:={inputs[k]}")
    assign_lines.append("")  # spacer
    for k, v in rows:
        if v is None:
            continue
        # sanitize key to a Mathcad-friendly identifier (best-effort)
        mc_key = k.replace(" ", "_").replace("-", "_").replace("/", "_").replace(".", "_")
        assign_lines.append(f"{mc_key}:={v}")
    p_txt.write_text("\n".join(assign_lines) + "\n", encoding="utf-8")

    return [p_json, p_in, p_out, p_txt]
