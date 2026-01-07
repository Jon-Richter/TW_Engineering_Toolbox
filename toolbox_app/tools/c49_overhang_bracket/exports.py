from __future__ import annotations

import csv
import json
from pathlib import Path
from typing import Any, Dict

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas

from .calc_trace import CalcTrace
from .report_renderer import render_report_html

def _autosize(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                val = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, len(val))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(80, max(10, max_len + 2))

def export_html(trace: CalcTrace, out_dir: Path) -> Path:
    html = render_report_html(trace)
    p = out_dir / "report.html"
    p.write_text(html, encoding="utf-8")
    return p

def export_calculation_report(trace: CalcTrace, out_dir: Path) -> Path:
    html = render_report_html(trace)
    p = out_dir / "calculation_report.html"
    p.write_text(html, encoding="utf-8")
    return p

def export_pdf(trace: CalcTrace, out_dir: Path) -> Path:
    """
    Minimal PDF: summary + hash/version. Full detail is in report.html.
    """
    p = out_dir / "report.pdf"
    c = canvas.Canvas(str(p), pagesize=letter)
    w, h = letter
    y = h - 72
    c.setFont("Helvetica-Bold", 14)
    c.drawString(72, y, "C49 Overhang Bracket – Calculation Package (Summary)")
    y -= 24
    c.setFont("Helvetica", 10)
    c.drawString(72, y, f"Tool: {trace.meta.tool_id} v{trace.meta.tool_version}")
    y -= 14
    c.drawString(72, y, f"Input hash: {trace.meta.input_hash}")
    y -= 14
    c.drawString(72, y, f"Generated: {trace.meta.timestamp}")
    y -= 22
    c.setFont("Helvetica", 9)
    c.drawString(72, y, "Note: Full step-by-step calcs are provided in report.html (offline).")
    y -= 18
    c.setFont("Helvetica-Bold", 10)
    c.drawString(72, y, "Key outputs:")
    y -= 14
    c.setFont("Helvetica", 9)
    for k, v in (trace.summary or {}).items():
        if y < 72:
            c.showPage()
            y = h - 72
            c.setFont("Helvetica", 9)
        c.drawString(84, y, f"{k}: {v}")
        y -= 12
    c.showPage()
    c.save()
    return p

def export_excel(trace: CalcTrace, out_dir: Path, results: Dict[str, Any]) -> Path:
    wb = Workbook()

    # Inputs sheet
    ws = wb.active
    ws.title = "Inputs"
    ws.append(["id", "label", "value", "units", "source", "notes"])
    for i in trace.inputs:
        ws.append([i.id, i.label, i.value, i.units, i.source, i.notes])
    _autosize(ws)

    # Assumptions
    ws2 = wb.create_sheet("Assumptions")
    ws2.append(["id", "text"])
    for a in trace.assumptions:
        ws2.append([a.id, a.text])
    _autosize(ws2)

    # Calcs
    ws3 = wb.create_sheet("Calcs")
    ws3.append(["id", "section", "title", "reference", "equation", "substitution", "result_rounded", "units", "variables_json"])
    for s in trace.steps:
        ref = "; ".join([f"{r.type}:{r.ref}" for r in s.references])
        vars_json = json.dumps([{
            "symbol": v.symbol, "description": v.description, "value": v.value, "units": v.units, "source": v.source
        } for v in s.variables], ensure_ascii=True)
        ws3.append([s.id, s.section, s.title, ref, s.equation_latex, s.substitution_latex, s.result_rounded.value, s.result_rounded.units, vars_json])
    _autosize(ws3)

    # Tables
    ws4 = wb.create_sheet("Tables")
    ws4.append(["table_name", "json"])
    for k, t in trace.tables.items():
        ws4.append([k, json.dumps(t, ensure_ascii=True)])
    _autosize(ws4)

    # Results
    ws5 = wb.create_sheet("Results")
    ws5.append(["key", "value"])
    for k, v in results.items():
        ws5.append([k, json.dumps(v, ensure_ascii=False) if isinstance(v, (dict, list)) else v])
    _autosize(ws5)

    p = out_dir / "results.xlsx"
    wb.save(p)
    return p

def export_json(trace: CalcTrace, out_dir: Path, results: Dict[str, Any]) -> Dict[str, Path]:
    p1 = out_dir / "calc_trace.json"
    p1.write_text(json.dumps(trace.to_dict(), indent=2, ensure_ascii=True), encoding="utf-8")

    p2 = out_dir / "results.json"
    p2.write_text(json.dumps(results, indent=2, ensure_ascii=True), encoding="utf-8")
    return {"calc_trace": p1, "results": p2}

def export_mathcad_handoff(trace: CalcTrace, out_dir: Path, results: Dict[str, Any]) -> Dict[str, Path]:
    # Inputs CSV
    p_csv = out_dir / "mathcad_inputs.csv"
    with p_csv.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["id", "label", "value", "units", "source", "notes"])
        for i in trace.inputs:
            w.writerow([i.id, i.label, i.value, i.units, i.source, i.notes])
        # include key results
        w.writerow([])
        w.writerow(["__RESULTS__", "", "", "", "", ""])
        for k, v in results.items():
            w.writerow([k, k, v, "-", "tool", ""])

    p_steps = out_dir / "mathcad_steps.json"
    steps = []
    for s in trace.steps:
        steps.append({
            "id": s.id,
            "section": s.section,
            "title": s.title,
            "output_symbol": s.output_symbol,
            "equation_latex": s.equation_latex,
            "substitution_latex": s.substitution_latex,
            "variables": [{
                "symbol": v.symbol, "description": v.description, "value": v.value, "units": v.units, "source": v.source
            } for v in s.variables],
            "result_unrounded": {"value": s.result_unrounded.value, "units": s.result_unrounded.units},
            "result_rounded": {"value": s.result_rounded.value, "units": s.result_rounded.units},
        })
    p_steps.write_text(json.dumps(steps, indent=2, ensure_ascii=True), encoding="utf-8")
    return {"mathcad_inputs": p_csv, "mathcad_steps": p_steps}

def export_all(trace: CalcTrace, out_dir: Path, results: Dict[str, Any]) -> Dict[str, Path]:
    outputs: Dict[str, Path] = {}
    outputs["html"] = export_html(trace, out_dir)
    outputs["calculation_report"] = export_calculation_report(trace, out_dir)
    outputs["pdf"] = export_pdf(trace, out_dir)
    outputs.update(export_json(trace, out_dir, results))
    outputs["excel"] = export_excel(trace, out_dir, results)
    outputs.update(export_mathcad_handoff(trace, out_dir, results))
    return outputs
