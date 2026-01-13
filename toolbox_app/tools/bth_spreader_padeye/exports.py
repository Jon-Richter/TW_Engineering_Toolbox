from __future__ import annotations
from pathlib import Path
from typing import Dict, Any
import json, csv
from .report_renderer import render_report_html

def export_all(run_dir: Path, trace: Dict[str, Any], results: Dict[str, Any]) -> Dict[str, str]:
    run_dir.mkdir(parents=True, exist_ok=True)

    (run_dir/"calc_trace.json").write_text(json.dumps(trace, indent=2), encoding="utf-8")
    (run_dir/"results.json").write_text(json.dumps(results, indent=2), encoding="utf-8")
    (run_dir/"report.html").write_text(render_report_html(trace), encoding="utf-8")

    from openpyxl import Workbook

    wb=Workbook()
    ws=wb.active; ws.title="Inputs"
    ws.append(["id","label","value","units","source","notes"])
    for i in trace["inputs"]:
        ws.append([i["id"], i["label"], i["value"], i["units"], i["source"], i.get("notes","")])
    wsA=wb.create_sheet("Assumptions"); wsA.append(["id","text"])
    for a in trace.get("assumptions",[]):
        wsA.append([a["id"], a["text"]])
    wsC=wb.create_sheet("Calcs")
    wsC.append(["id","section","title","reference","equation","substitution","result_rounded","units","variables_json"])
    for s in trace["steps"]:
        refs="; ".join([f"{r['type']}:{r['ref']}" for r in s.get("references",[])])
        wsC.append([s["id"], s["section"], s["title"], refs, s["equation_latex"], s["substitution_latex"],
                    s["result_rounded"]["value"], s["result_rounded"]["units"], json.dumps(s["variables"])])
    wsT=wb.create_sheet("Tables"); wsT.append(["name","json"])
    for k,v in (trace.get("tables") or {}).items():
        wsT.append([k, json.dumps(v)])
    xlsx=run_dir/"design.xlsx"; wb.save(xlsx)

    with (run_dir/"mathcad_inputs.csv").open("w", newline="", encoding="utf-8") as f:
        wcsv=csv.writer(f); wcsv.writerow(["id","label","value","units","source"])
        for i in trace["inputs"]:
            wcsv.writerow([i["id"], i["label"], i["value"], i["units"], i["source"]])
        for k,v in (results.get("key_outputs") or {}).items():
            if isinstance(v, dict) and "value" in v:
                wcsv.writerow([f"out:{k}", k, v["value"], v.get("units",""), "results"])

    (run_dir/"mathcad_steps.json").write_text(json.dumps(trace["steps"], indent=2), encoding="utf-8")

    return {
        "report.html": str(run_dir/"report.html"),
        "calc_trace.json": str(run_dir/"calc_trace.json"),
        "results.json": str(run_dir/"results.json"),
        "design.xlsx": str(xlsx),
        "mathcad_inputs.csv": str(run_dir/"mathcad_inputs.csv"),
        "mathcad_steps.json": str(run_dir/"mathcad_steps.json"),
    }
